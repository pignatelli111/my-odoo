# -*- coding: utf-8 -*-
"""Import ANACO / Voci Contrattuali_SAL from ANACO_REV7-style .xlsx (openpyxl, data_only)."""
import base64
import io
import math
import re
from odoo import api, fields, models, _
from odoo.exceptions import UserError

try:
    import openpyxl
except ImportError:  # pragma: no cover
    openpyxl = None

from ..models.sbu_cost_family import (
    infer_cost_family_from_pos,
    infer_cost_family_from_price_cost_vals,
)


# Column indices (1-based) validated against ANACO_REV7_111122.xlsx layout.
ANACO_ROW_PARAMS = 5  # sheet-level Sc multipliers K5:M5, cost % BM/BP, …
ANACO_ROW_FIRST_DATA_DEFAULT = 12

ANACO_COL_POS = 2  # B
ANACO_COL_DESC = 3  # C
ANACO_COL_B_MM = 4  # D
ANACO_COL_H_MM = 6  # F
ANACO_COL_QTY = 8  # H
ANACO_COL_SC_MULT = (11, 12, 13)  # K, L, M — row 5 multipliers, not row data discounts
ANACO_COL_PRICE = {
    'price_serramento_cad': 14,
    'price_oscuramento_cad': 28,
    'price_accessori_cad': 22,
    'price_kit_avvolgimento_cad': 32,
    'price_cassonetto_cad': 34,
    'price_automatismo_cad': 36,
    'price_zanzariera_cad': 38,
    'price_vetro_cad': 40,
    'price_pannello_cad': 42,
    'price_controtelaio_cad': 44,
    'price_trasformazione_cad': 46,
    'price_smontaggio_cad': 48,
    'price_nolo_cad': 52,
}
ANACO_COL_COST_COIB = 30
ANACO_COL_COST_POSA_LIN = 50
# Col 52 is price nolo (ANACO_COL_PRICE); no separate cost_nolo field on Odoo lines.
ANACO_COL_COST_TRASPORTO = 57
ANACO_COL_COST_TECH_PM = 59
ANACO_COL_COST_CANTIERE = 61
ANACO_COL_COST_EXTRA = 63
ANACO_COL_COST_IND_PCT = 65  # BM
ANACO_COL_COST_MOL_PCT = 68  # BP
ANACO_COL_BS_UNIT = 71
ANACO_COL_NOTE = 74

SAL_ROW_FIRST_DATA_DEFAULT = 16
SAL_COL_ITEM = 2
SAL_COL_DESC = 3
SAL_COL_QTY = 4
SAL_COL_TOT_ML = 5
SAL_COL_MQ = 6
SAL_COL_TOT_MQ = 7
SAL_COL_UNIT_PRICE = 8
SAL_COL_FLOOR_BLOCKS = (
    (12, 15),  # PT: L–O
    (16, 19),
    (20, 23),
    (24, 27),
    (28, 31),
    (32, 35),
    (36, 39),
    (40, 43),
    (44, 47),
)
SAL_FLOOR_FIELDS = (
    'floor_pt', 'floor_p1', 'floor_p2', 'floor_p3', 'floor_p4',
    'floor_p5', 'floor_p6', 'floor_p7', 'floor_p8',
)
SAL_COL_SAL_START = 98  # SAL-1 … SAL-10

# OFFERTA sheet (client-facing offer grid) — fallback when ANACO B/C are empty
OFFERTA_ROW_FIRST_DATA_DEFAULT = 24
OFFERTA_COL_CODE = 2
OFFERTA_COL_WIDTH = 3
OFFERTA_COL_HEIGHT = 4
OFFERTA_COL_DESC = 5
OFFERTA_COL_UOM = 6
OFFERTA_COL_QTY = 7
OFFERTA_COL_PRICE = 8

_ANACO_SKIP_B = frozenset({
    'POS.', 'COD.', 'ITEMS', 'DATA', 'RIFERIMENTO', 'OGGETTO', 'CLIENTE',
})
_SAL_SKIP_ITEM = frozenset({
    'ITEM', 'CODICE', 'COD.', 'POS.', 'DATA', '0', '0.0',
})
_ANACO_PRODUCT_POS_RE = re.compile(
    r'^[A-Z]{1,6}\d',  # F1, F3a, LA01, ACC1, …
    re.IGNORECASE,
)
_ANACO_NUMERIC_COLS = tuple(
    set(ANACO_COL_PRICE.values())
    | {
        ANACO_COL_B_MM, ANACO_COL_H_MM, ANACO_COL_QTY, ANACO_COL_BS_UNIT,
        ANACO_COL_COST_COIB, ANACO_COL_COST_POSA_LIN, ANACO_COL_COST_TRASPORTO,
        ANACO_COL_COST_TECH_PM, ANACO_COL_COST_CANTIERE, ANACO_COL_COST_EXTRA,
    }
)


def _sheet_by_aliases(wb, aliases):
    lower = {n.lower().replace(' ', '_'): n for n in wb.sheetnames}
    for a in aliases:
        key = a.lower().replace(' ', '_')
        if key in lower:
            return wb[lower[key]]
    return None


def _filter_model_fields(model, vals):
    allowed = set(model._fields)
    return {k: v for k, v in vals.items() if k in allowed}


def _anaco_pos_looks_like_product(pos_raw):
    """P1002 / REV7: F1, F3a, LA01 — not COD., ITEMS, or section titles without code."""
    pos = (pos_raw or '').strip()
    if not pos:
        return False
    pos_up = pos.upper()
    if pos_up in _ANACO_SKIP_B or pos_up.startswith('CLIENTE'):
        return False
    first_line = pos.split('\n')[0].strip()
    return bool(_ANACO_PRODUCT_POS_RE.match(first_line))


def _anaco_row_is_data(sh, row):
    """True for product rows (qty/BS/dimensions), not testata or section titles."""
    pos = _cell_str(sh, row, ANACO_COL_POS)
    if _anaco_pos_looks_like_product(pos):
        return True
    b_mm = _cell_num(sh, row, ANACO_COL_B_MM)
    h_mm = _cell_num(sh, row, ANACO_COL_H_MM)
    if (b_mm and b_mm > 0) and (h_mm and h_mm > 0):
        return True
    qty = _cell_num(sh, row, ANACO_COL_QTY) or 0.0
    bs = _cell_num(sh, row, ANACO_COL_BS_UNIT) or 0.0
    if qty > 0 and bs > 0:
        return True
    if _anaco_pos_looks_like_product(pos) and (qty > 0 or bs > 0):
        return True
    return False


def _detect_anaco_first_row(sh, fallback):
    for row in range(1, 81):
        if _anaco_row_is_data(sh, row):
            return row
    return fallback


def _sal_row_is_importable(sh_vals, sh_form, row):
    """Contractual SAL line with item code and billable amount (P1002: from row 17)."""
    item = _cell_str_merged(sh_vals, sh_form, row, SAL_COL_ITEM).strip()
    desc = _cell_str_merged(sh_vals, sh_form, row, SAL_COL_DESC).strip()
    if item.upper() in _SAL_SKIP_ITEM:
        return False
    if desc.upper() in ("DESCRIZIONE D'OFFERTA", 'DESCRIZIONE'):
        return False
    unit = _cell_num_merged(sh_vals, sh_form, row, SAL_COL_UNIT_PRICE) or 0.0
    qty = _cell_num_merged(sh_vals, sh_form, row, SAL_COL_QTY) or 0.0
    tot_mq = _cell_num_merged(sh_vals, sh_form, row, SAL_COL_TOT_MQ) or 0.0
    if not item or item in ('0', '0.0'):
        return False
    return unit > 0 or qty > 0 or tot_mq > 0


def _detect_sal_first_row(sh_vals, sh_form, fallback):
    sh_form = sh_form or sh_vals
    for row in range(1, 81):
        if _sal_row_is_importable(sh_vals, sh_form, row):
            return row
    return fallback


def _get_workbook_sheet_pair(raw):
    """Return (values_sheet, formula_sheet) for ANACO — merge literals from formula book."""
    buf = io.BytesIO(raw)
    wb_vals = openpyxl.load_workbook(buf, data_only=True, read_only=False)
    buf.seek(0)
    wb_form = openpyxl.load_workbook(buf, data_only=False, read_only=False)
    return wb_vals, wb_form


def _cell_num_merged(sh_vals, sh_form, row, col):
    v = _cell_num(sh_vals, row, col)
    if v is not None and v != 0:
        return v
    raw = sh_form.cell(row, col).value
    if raw is None or raw == '':
        return None
    if isinstance(raw, str) and raw.strip().startswith('='):
        return None
    return _cell_num(sh_form, row, col)


def _cell_str_merged(sh_vals, sh_form, row, col):
    s = _cell_str(sh_vals, row, col)
    if s:
        return s
    raw = sh_form.cell(row, col).value
    if raw is None or raw == '':
        return ''
    if isinstance(raw, str) and raw.strip().startswith('='):
        return ''
    return _cell_str(sh_form, row, col)


def _cell_str(sh, row, col):
    v = sh.cell(row, col).value
    if v is None:
        return ''
    if isinstance(v, (int, float)):
        if v == 0:
            return ''
        if float(v).is_integer():
            return str(int(v))
        return str(v).strip()
    return str(v).strip()


def _cell_num(sh, row, col):
    v = sh.cell(row, col).value
    if v is None or v == '':
        return None
    if isinstance(v, str):
        vs = v.replace(',', '.').strip()
        if not vs or vs == '-' or vs.upper().startswith('#'):
            return None
        v = vs
    try:
        x = float(v)
    except (TypeError, ValueError):
        return None
    if isinstance(x, float) and (math.isnan(x) or math.isinf(x)):
        return None
    return x


def _sal_header_normalize(text):
    """Compact + raw upper labels for SAL-n header detection."""
    raw = (text or '').strip().upper()
    compact = re.sub(r'[\s._\-/]+', '', raw)
    return compact, raw


def _sal_header_is_sal_n(compact, n):
    return bool(re.fullmatch(rf'SAL0?{n}', compact))


# Floor grid ~cols 12–47; SAL % headers may start ~48 (client) or 98 (REV7).
SAL_PCT_HEADER_SCAN_MIN = 48
SAL_PCT_DATA_SCAN_MIN = 85


def _detect_sal_pct_start_column(sh_vals, sh_form=None, fallback=SAL_COL_SAL_START):
    """
    Find 1-based column of SAL-1 % in the SAL sheet header row.
    REV7 uses column 98; client workbooks (e.g. P1002 REV03) often shift left.
    """
    sheets = [sh_vals]
    if sh_form is not None and sh_form is not sh_vals:
        sheets.append(sh_form)

    scan_from = SAL_PCT_HEADER_SCAN_MIN
    for sh in sheets:
        max_col = min((sh.max_column or 150) + 1, 220)
        # Full sequence SAL-1 … SAL-10 on one header row
        for row in range(1, 40):
            for start in range(scan_from, max_col - 8):
                if all(
                    _sal_header_is_sal_n(
                        _sal_header_normalize(_cell_str(sh, row, start + i))[0],
                        i + 1,
                    )
                    for i in range(10)
                ):
                    return start
        for row in range(1, 40):
            for col in range(scan_from, max_col):
                compact, raw = _sal_header_normalize(_cell_str(sh, row, col))
                is_sal1 = (
                    _sal_header_is_sal_n(compact, 1)
                    or bool(re.search(r'SAL\s*[-.]?\s*1(?!\d)', raw))
                    or compact in ('SAL1PCT', 'SAL1PERC', 'PERCSAL1')
                )
                if not is_sal1:
                    continue
                for off in range(1, 6):
                    c2, _raw2 = _sal_header_normalize(_cell_str(sh, row, col + off))
                    if _sal_header_is_sal_n(c2, 2):
                        return col
                if _sal_header_is_sal_n(compact, 1):
                    return col
    return fallback


def _coerce_import_sal_pct(raw):
    """Return a SAL period % in 0–100, or None if the cell looks like €/mq/etc."""
    if raw is None:
        return None
    pct = _norm_pct(raw)
    if pct <= 0.0 or pct > 100.0:
        return None
    return pct


def _sal_pct_block_row_pcts(sh_vals, sh_form, row, start_col):
    """List of coerced SAL period % (0–100) for 10 consecutive columns."""
    pcts = []
    for i in range(10):
        v = _cell_num_merged(sh_vals, sh_form, row, start_col + i)
        pct = _coerce_import_sal_pct(v)
        if pct is not None:
            pcts.append(pct)
    return pcts


def _sal_pct_plan_looks_incremental(pcts):
    """
    True when cells look like SAL-1…10 planning quotas (not one «100%» cumulative cell).
    A single 100 in one column is usually «avanzamento cumulativo», not period split.
    """
    if not pcts:
        return False
    if len(pcts) == 1 and pcts[0] >= 99.9:
        return False
    return sum(pcts) <= 100.0000001


def _sal_pct_block_row_sum(sh_vals, sh_form, row, start_col):
    """Sum of incremental SAL period %; None if block does not look like SAL planning."""
    pcts = _sal_pct_block_row_pcts(sh_vals, sh_form, row, start_col)
    if not _sal_pct_plan_looks_incremental(pcts):
        return None
    return sum(pcts)


def _detect_sal_pct_by_data_profile(sh_vals, sh_form, first_row, fallback):
    """
    If headers are missing/merged, pick the 10-column block where data rows
    look like SAL period % (each 0–100, row sum <= 100).
    """
    sh_form = sh_form or sh_vals
    max_col = min((sh_vals.max_column or 150) + 1, 220)
    best_col = None
    best_score = 0
    sample_rows = list(range(first_row, min(first_row + 25, (sh_vals.max_row or first_row) + 1)))
    if not sample_rows:
        return fallback

    scan_from = max(SAL_PCT_DATA_SCAN_MIN, SAL_COL_SAL_START - 15)
    for start in range(scan_from, max_col - 9):
        score = 0
        for r in sample_rows:
            item = _cell_str_merged(sh_vals, sh_form, r, SAL_COL_ITEM)
            desc = _cell_str_merged(sh_vals, sh_form, r, SAL_COL_DESC)
            if not item and not desc:
                continue
            if _sal_pct_block_row_sum(sh_vals, sh_form, r, start) is not None:
                score += 1
        if score > best_score:
            best_score = score
            best_col = start
    if best_score >= 2:
        return best_col
    return fallback


def _norm_pct(val):
    """Excel may store 5 as 5% or 0.05 as 5%."""
    if val is None:
        return 0.0
    try:
        v = float(val)
    except (TypeError, ValueError):
        return 0.0
    if 0 < v <= 1.0:
        return v * 100.0
    return v


def _mult_to_discount_pct(mult):
    """K5/L5/M5 style multiplier 1 = 0% discount, 0.85 = 15%."""
    if mult is None:
        return 0.0
    try:
        m = float(mult)
    except (TypeError, ValueError):
        return 0.0
    if m >= 1.0 or m <= 0.0:
        return 0.0
    return max(0.0, min(100.0, (1.0 - m) * 100.0))


class SbuEstimateAnacoImportWizard(models.TransientModel):
    _name = 'sbu.estimate.anaco.import.wizard'
    _description = 'Import ANACO / SAL from Excel'

    estimate_id = fields.Many2one(
        'sbu.estimate',
        string='Preventivo',
        required=True,
        ondelete='cascade',
    )
    data_file = fields.Binary(string='File .xlsx')
    data_filename = fields.Char(string='Nome file')

    import_anaco = fields.Boolean(string='Importa righe ANACO', default=True)
    import_sal = fields.Boolean(string='Importa Voci Contrattuali SAL', default=True)
    replace_anaco_lines = fields.Boolean(
        string='Sostituisci righe preventivo esistenti',
        default=True,
        help='Se attivo, elimina le righe ANACO (sbu.estimate.line) prima dell\'import.',
    )
    replace_sal_lines = fields.Boolean(
        string='Sostituisci righe SAL esistenti',
        default=True,
    )
    anaco_first_row = fields.Integer(
        string='Prima riga dati ANACO (1-based)',
        default=ANACO_ROW_FIRST_DATA_DEFAULT,
        required=True,
        help='Default 12 for REV7. If no rows are found, the wizard scans upward from row 1 '
             'when auto-detect is enabled.',
    )
    auto_detect_first_row = fields.Boolean(
        string='Rileva automaticamente prima riga ANACO',
        default=True,
    )
    import_offerta_fallback = fields.Boolean(
        string='Se ANACO vuoto, importa da foglio OFFERTA',
        default=True,
        help='Many workbooks only fill the OFFERTA grid (CODICE / DESCRIZIONE / PREZZO) '
             'while ANACO columns B–C stay empty.',
    )
    sal_first_row = fields.Integer(
        string='Prima riga dati SAL (1-based)',
        default=SAL_ROW_FIRST_DATA_DEFAULT,
        required=True,
        help='REV7 default 16; P1002 REV03 often starts at 17 (after section title row).',
    )
    auto_detect_sal_first_row = fields.Boolean(
        string='Rileva automaticamente prima riga SAL',
        default=True,
    )
    auto_detect_sal_columns = fields.Boolean(
        string='Rileva colonna SAL-1 automaticamente',
        default=True,
        help='Cerca l’intestazione SAL-1 nel foglio SAL (REV7 = colonna 98; '
             'file cliente possono usare colonne diverse).',
    )
    sal_col_sal_start = fields.Integer(
        string='Colonna SAL-1 (1-based)',
        default=SAL_COL_SAL_START,
        help='Usata solo se «Rileva colonna SAL-1» è disattivato. REV7 = 98.',
    )
    default_commission_pct = fields.Float(
        string='Comm. % default (righe ANACO)',
        default=0.0,
        digits=(16, 2),
        help='La colonna «Comm.» in testata ANACO non è mappata per-riga in REV7; '
             'impostare qui la commissione da applicare a tutte le righe importate, se serve.',
    )

    @api.model
    def default_get(self, fields_list):
        res = super().default_get(fields_list)
        if not res.get('estimate_id') and self.env.context.get('default_estimate_id'):
            res['estimate_id'] = self.env.context['default_estimate_id']
        return res

    def _sbu_validate_upload(self):
        self.ensure_one()
        if not self.data_file:
            raise UserError(_('Selezionare un file .xlsx.'))
        name = (self.data_filename or '').lower()
        if name and not name.endswith(('.xlsx', '.xlsm')):
            raise UserError(
                _('Formato non supportato: %(name)s. Salvare il file Excel come .xlsx (non .xls).',
                  name=self.data_filename)
            )

    def _import_anaco_sheet_rows(self, estimate, anaco_sh, anaco_sh_form, Line, first_row):
        """Import ANACO sheet rows; returns (count, optional chatter note)."""
        sh_form = anaco_sh_form or anaco_sh

        def _num(row, col):
            return _cell_num_merged(anaco_sh, sh_form, row, col)

        def _str(row, col):
            return _cell_str_merged(anaco_sh, sh_form, row, col)

        sc1 = _mult_to_discount_pct(_num(ANACO_ROW_PARAMS, ANACO_COL_SC_MULT[0]))
        sc2 = _mult_to_discount_pct(_num(ANACO_ROW_PARAMS, ANACO_COL_SC_MULT[1]))
        sc3 = _mult_to_discount_pct(_num(ANACO_ROW_PARAMS, ANACO_COL_SC_MULT[2]))
        ind_pct = _norm_pct(_num(ANACO_ROW_PARAMS, ANACO_COL_COST_IND_PCT))
        mol_pct = _norm_pct(_num(ANACO_ROW_PARAMS, ANACO_COL_COST_MOL_PCT))

        seq = 10
        empty_run = 0
        count = 0
        max_row = anaco_sh.max_row or first_row
        note = False

        for r in range(first_row, max_row + 1):
            if not _anaco_row_is_data(anaco_sh, r) and not _anaco_row_is_data(sh_form, r):
                empty_run += 1
                if empty_run >= 8:
                    break
                continue
            empty_run = 0

            pos = _str(r, ANACO_COL_POS)
            desc = _str(r, ANACO_COL_DESC)
            if not desc:
                if pos:
                    desc = pos
                else:
                    b_mm = _num(r, ANACO_COL_B_MM) or 0.0
                    h_mm = _num(r, ANACO_COL_H_MM) or 0.0
                    if b_mm and h_mm:
                        desc = _('Item %(w)s×%(h)s mm') % {'w': int(b_mm), 'h': int(h_mm)}
                    else:
                        desc = _('Riga Excel %(row)s') % {'row': r}

            b_mm = _num(r, ANACO_COL_B_MM) or 0.0
            h_mm = _num(r, ANACO_COL_H_MM) or 0.0
            calc_uom = 'mq' if b_mm and h_mm else 'nr'

            line_vals = {
                'estimate_id': estimate.id,
                'sequence': seq,
                'pos': pos or False,
                'item_code': pos or False,
                'description': desc,
                'calc_uom_type': calc_uom,
                'qty': _num(r, ANACO_COL_QTY) or 1.0,
                'width_mm': b_mm,
                'height_mm': h_mm,
                'discount_sc1': sc1,
                'discount_sc2': sc2,
                'discount_sc3': sc3,
                'commission_pct': self.default_commission_pct or 0.0,
                'cost_industrial_pct': ind_pct,
                'cost_mol_pct': mol_pct,
            }
            note_parts = []
            for fname, col in ANACO_COL_PRICE.items():
                v = _num(r, col)
                if v is not None:
                    line_vals[fname] = v
            for fname, col in (
                ('cost_coibentazione_cad', ANACO_COL_COST_COIB),
                ('cost_posa_lamiera_lin_cad', ANACO_COL_COST_POSA_LIN),
                ('cost_trasporto_cad', ANACO_COL_COST_TRASPORTO),
                ('cost_tech_pm_cad', ANACO_COL_COST_TECH_PM),
                ('cost_cantiere_cad', ANACO_COL_COST_CANTIERE),
                ('cost_extra_cad', ANACO_COL_COST_EXTRA),
            ):
                v = _num(r, col)
                if v is not None:
                    line_vals[fname] = v

            bs = _num(r, ANACO_COL_BS_UNIT)
            if bs is not None:
                line_vals['price_anaco_bs_cad'] = bs

            ntxt = _str(r, ANACO_COL_NOTE)
            if ntxt:
                note_parts.append(ntxt)
            if note_parts:
                line_vals['note'] = '\n'.join(note_parts)

            cost_family = infer_cost_family_from_pos(pos) or infer_cost_family_from_price_cost_vals(line_vals)
            if cost_family:
                line_vals['cost_family'] = cost_family

            try:
                Line.create(_filter_model_fields(Line, line_vals))
            except Exception as err:  # pylint: disable=broad-except
                raise UserError(
                    _('Errore import riga ANACO Excel %(row)s (pos. %(pos)s): %(err)s')
                    % {'row': r, 'pos': pos or '—', 'err': err}
                ) from err
            count += 1
            seq += 10

        if count and first_row != self.anaco_first_row:
            note = _('Import ANACO: prima riga dati rilevata alla riga %(row)s.', row=first_row)
        return count, note

    def _import_offerta_sheet_rows(self, estimate, offerta_sh, offerta_form, Line):
        """Import client offer grid when ANACO identity columns are empty."""
        sh_form = offerta_form or offerta_sh

        def _num(row, col):
            return _cell_num_merged(offerta_sh, sh_form, row, col)

        def _str(row, col):
            return _cell_str_merged(offerta_sh, sh_form, row, col)

        first_row = OFFERTA_ROW_FIRST_DATA_DEFAULT
        for row in range(20, 41):
            if _str(row, OFFERTA_COL_CODE).upper() == 'CODICE':
                first_row = row + 1
                break

        seq = 10
        count = 0
        empty_run = 0
        max_row = offerta_sh.max_row or first_row

        for r in range(first_row, max_row + 1):
            code = _str(r, OFFERTA_COL_CODE)
            desc = _str(r, OFFERTA_COL_DESC)
            qty = _num(r, OFFERTA_COL_QTY)
            price = _num(r, OFFERTA_COL_PRICE)
            if code.upper() in ('CODICE', 'ITEM', '0'):
                continue
            if not code and not desc and not (qty and qty > 0) and not (price and price > 0):
                empty_run += 1
                if empty_run >= 8:
                    break
                continue
            empty_run = 0

            b_mm = _num(r, OFFERTA_COL_WIDTH) or 0.0
            h_mm = _num(r, OFFERTA_COL_HEIGHT) or 0.0
            uom_raw = (_str(r, OFFERTA_COL_UOM) or '').lower()
            if 'mq' in uom_raw or 'm²' in uom_raw:
                calc_uom = 'mq'
            elif 'ml' in uom_raw or 'm.l' in uom_raw:
                calc_uom = 'ml'
            elif 'corpo' in uom_raw or 'a corpo' in uom_raw:
                calc_uom = 'corpo'
            elif b_mm and h_mm:
                calc_uom = 'mq'
            else:
                calc_uom = 'nr'

            if not desc:
                desc = code or _('Voce OFFERTA riga %(row)s') % {'row': r}

            line_vals = {
                'estimate_id': estimate.id,
                'sequence': seq,
                'pos': code or False,
                'item_code': code or False,
                'description': desc,
                'calc_uom_type': calc_uom,
                'qty': qty or 1.0,
                'width_mm': b_mm,
                'height_mm': h_mm,
                'commission_pct': self.default_commission_pct or 0.0,
            }
            if price and price > 0:
                line_vals['price_anaco_bs_cad'] = price

            cost_family = infer_cost_family_from_pos(code) or infer_cost_family_from_price_cost_vals(line_vals)
            if cost_family:
                line_vals['cost_family'] = cost_family

            Line.create(_filter_model_fields(Line, line_vals))
            count += 1
            seq += 10
        return count

    def action_import(self):
        if not openpyxl:
            raise UserError(_('Installare la libreria Python openpyxl sul server Odoo.'))
        self.ensure_one()
        if not self.import_anaco and not self.import_sal:
            raise UserError(_('Selezionare almeno un tipo di import (ANACO e/o SAL).'))
        self._sbu_validate_upload()

        raw = base64.b64decode(self.data_file)
        try:
            wb_vals, wb_form = _get_workbook_sheet_pair(raw)
        except Exception as err:  # pylint: disable=broad-except
            raise UserError(_('Impossibile leggere il file Excel: %s') % (err,)) from err

        anaco_sh = _sheet_by_aliases(wb_vals, ('ANACO', 'Anaco'))
        anaco_sh_form = _sheet_by_aliases(wb_form, ('ANACO', 'Anaco')) if wb_form else None
        sal_sh = _sheet_by_aliases(wb_vals, (
            'Voci Contrattuali_SAL',
            'Voci Contrattuali SAL',
            'VOCI CONTRATTUALI_SAL',
        ))
        sal_sh_form = _sheet_by_aliases(wb_form, (
            'Voci Contrattuali_SAL',
            'Voci Contrattuali SAL',
            'VOCI CONTRATTUALI_SAL',
        )) if wb_form else None

        n_anaco = 0
        n_sal = 0
        sal_pct_col_used = None
        estimate = self.estimate_id

        Line = self.env['sbu.estimate.line']
        SalLine = self.env['sbu.estimate.sal.line']

        if self.import_anaco:
            if not anaco_sh:
                raise UserError(
                    _('Foglio «ANACO» non trovato. Fogli presenti: %s')
                    % ', '.join(wb_vals.sheetnames)
                )
            if self.replace_anaco_lines:
                estimate.line_ids.unlink()

            first_row = self.anaco_first_row
            if self.auto_detect_first_row:
                detected = _detect_anaco_first_row(anaco_sh, first_row)
                if detected != first_row:
                    first_row = detected

            n_anaco, import_note = self._import_anaco_sheet_rows(
                estimate, anaco_sh, anaco_sh_form, Line, first_row,
            )

            if n_anaco == 0 and self.auto_detect_first_row and first_row != 1:
                n_anaco, retry_note = self._import_anaco_sheet_rows(
                    estimate, anaco_sh, anaco_sh_form, Line, 1,
                )
                if retry_note:
                    import_note = retry_note

            if n_anaco == 0 and self.import_offerta_fallback:
                offerta_sh = _sheet_by_aliases(wb_vals, ('OFFERTA', 'Offerta'))
                offerta_form = _sheet_by_aliases(wb_form, ('OFFERTA', 'Offerta'))
                if offerta_sh:
                    n_offerta = self._import_offerta_sheet_rows(
                        estimate, offerta_sh, offerta_form, Line,
                    )
                    if n_offerta:
                        n_anaco = n_offerta
                        import_note = _('Importate %(n)d righe dal foglio OFFERTA.') % {'n': n_offerta}

            if n_anaco == 0:
                raise UserError(
                    _('Nessuna riga importata dal file «%(file)s».\n\n'
                      'Controllare:\n'
                      '• Il foglio ANACO ha Pos./Descrizione (col. B–C) oppure dimensioni/prezzi (D, F, H, BS…)\n'
                      '• Oppure compilare il foglio OFFERTA (CODICE / DESCRIZIONE / PREZZO)\n'
                      '• Aprire il file in Excel e salvarlo di nuovo (valori calcolati)\n'
                      '• Prima riga dati: provare %(row)s o attivare «Rileva automaticamente»\n\n'
                      'Fogli trovati: %(sheets)s')
                    % {
                        'file': self.data_filename or 'xlsx',
                        'row': self.anaco_first_row,
                        'sheets': ', '.join(wb_vals.sheetnames),
                    }
                )
            if import_note:
                estimate.message_post(body=import_note)

        if self.import_sal:
            if not sal_sh:
                raise UserError(
                    _('Foglio «Voci Contrattuali_SAL» non trovato. Fogli presenti: %s')
                    % ', '.join(wb_vals.sheetnames)
                )
            if self.replace_sal_lines:
                estimate.sal_line_ids.unlink()

            sal_form = sal_sh_form or sal_sh

            def _sal_num(row, col):
                return _cell_num_merged(sal_sh, sal_form, row, col)

            def _sal_str(row, col):
                return _cell_str_merged(sal_sh, sal_form, row, col)

            sal_first_row = self.sal_first_row
            if self.auto_detect_sal_first_row:
                detected_sal_row = _detect_sal_first_row(sal_sh, sal_form, sal_first_row)
                if detected_sal_row != sal_first_row:
                    sal_first_row = detected_sal_row

            sal_pct_col = SAL_COL_SAL_START
            if self.auto_detect_sal_columns:
                sal_pct_col = _detect_sal_pct_start_column(
                    sal_sh, sal_sh_form, SAL_COL_SAL_START,
                )
                if sal_pct_col == SAL_COL_SAL_START:
                    sal_pct_col = _detect_sal_pct_by_data_profile(
                        sal_sh, sal_form, sal_first_row, SAL_COL_SAL_START,
                    )
            else:
                sal_pct_col = self.sal_col_sal_start or SAL_COL_SAL_START
            sal_pct_col_used = sal_pct_col

            seq = 10
            empty_run = 0
            max_row = sal_sh.max_row or sal_first_row
            sal_lines_with_pct = 0
            sal_pct_skipped_rows = 0
            for r in range(sal_first_row, max_row + 1):
                if not _sal_row_is_importable(sal_sh, sal_form, r):
                    empty_run += 1
                    if empty_run >= 5:
                        break
                    continue
                empty_run = 0
                item = _sal_str(r, SAL_COL_ITEM)
                desc = _sal_str(r, SAL_COL_DESC)
                if not desc:
                    desc = item or _('Voce SAL importata')

                qty = _sal_num(r, SAL_COL_QTY)
                unit = _sal_num(r, SAL_COL_UNIT_PRICE)
                tot_ml = _sal_num(r, SAL_COL_TOT_ML) or 0.0
                mq = _sal_num(r, SAL_COL_MQ) or 0.0
                tot_mq = _sal_num(r, SAL_COL_TOT_MQ) or 0.0

                if tot_ml and tot_ml > 0 and not (mq or tot_mq):
                    uom = 'ml'
                elif mq or tot_mq:
                    uom = 'mq'
                else:
                    uom = 'nr'

                sal_vals = {
                    'estimate_id': estimate.id,
                    'sequence': seq,
                    'item_ref': item or False,
                    'description': desc,
                    'uom_type': uom,
                    'qty_contract': qty or 0.0,
                    'unit_price': unit or 0.0,
                    'retention_percent': self.env['sbu.estimate.sal.line']._sbu_default_retention_percent(),
                }
                for (c0, c1), fname in zip(SAL_COL_FLOOR_BLOCKS, SAL_FLOOR_FIELDS):
                    block_sum = 0.0
                    for c in range(c0, c1 + 1):
                        v = _sal_num(r, c)
                        if v is not None:
                            block_sum += v
                    if block_sum:
                        sal_vals[fname] = block_sum

                pct_by_field = {}
                for i in range(10):
                    v = _sal_num(r, sal_pct_col + i)
                    pct = _coerce_import_sal_pct(v)
                    if pct is not None:
                        pct_by_field[f'sal_{i + 1}_pct'] = pct
                pct_values = list(pct_by_field.values())
                if not _sal_pct_plan_looks_incremental(pct_values):
                    sal_pct_skipped_rows += 1
                else:
                    sal_vals.update(pct_by_field)
                    sal_lines_with_pct += 1

                try:
                    SalLine.create(_filter_model_fields(SalLine, sal_vals))
                except Exception as err:  # pylint: disable=broad-except
                    raise UserError(
                        _('Errore import riga SAL Excel %(row)s: %(err)s') % {'row': r, 'err': err}
                    ) from err
                n_sal += 1
                seq += 10

            if sal_pct_skipped_rows:
                estimate.message_post(body=_(
                    'Import SAL: %(n)d righe senza %% SAL-1…10 importate '
                    '(colonna %(col)d: spesso «100%%» cumulativo o importi al posto '
                    'delle quote per periodo). Controllare in Excel le colonne SAL-1…10 '
                    'e, se serve, disattivare «Rileva colonna SAL-1» e indicare la colonna corretta.'
                ) % {'n': sal_pct_skipped_rows, 'col': sal_pct_col})
            elif n_sal and sal_lines_with_pct:
                pct_100 = estimate.sal_line_ids.filtered(
                    lambda l: (l.cumulative_pct or 0) >= 99.9
                )
                if len(pct_100) > len(estimate.sal_line_ids) * 0.8:
                    estimate.message_post(body=_(
                        'Import SAL: molte voci hanno Cum.%% ≈ 100%%. '
                        'Verificare che le colonne SAL-1…10 in Excel siano le quote '
                        'per ogni SAL (es. 20+30+50), non la colonna «avanzamento cumulativo».'
                    ))
            if n_sal and not sal_lines_with_pct:
                estimate.message_post(body=_(
                    'Import SAL: nessuna percentuale SAL-1…10 nel file '
                    '(colonna %(col)d; in P1002 REV03 le celle SAL-1…10 sono spesso '
                    'vuote — compilare le %% in Odoo o nel foglio prima del re-import). '
                    'Prima riga dati SAL usata: %(row)d.'
                ) % {'col': sal_pct_col, 'row': sal_first_row})
            elif sal_first_row != self.sal_first_row:
                estimate.message_post(body=_(
                    'Import SAL: prima riga dati rilevata alla riga Excel %(row)d '
                    '(default %(def)s; es. titolo sezione «SERRAMENTI SERIE F» alla riga 16).'
                ) % {'row': sal_first_row, 'def': self.sal_first_row})

        wb_vals.close()
        wb_form.close()

        body = _(
            'Import Excel completato: %(anaco)d righe ANACO, %(sal)d righe SAL.'
        ) % {'anaco': n_anaco, 'sal': n_sal}
        if self.import_sal and sal_pct_col_used:
            body += '<br/>' + _(
                'Colonne SAL-1…10: colonna Excel %(col)d (REV7 predefinita: %(rev7)d).'
            ) % {'col': sal_pct_col_used, 'rev7': SAL_COL_SAL_START}
        estimate.message_post(body=body)
        return {
            'type': 'ir.actions.act_window',
            'name': _('Preventivo'),
            'res_model': 'sbu.estimate',
            'res_id': estimate.id,
            'view_mode': 'form',
            'views': [(False, 'form')],
            'target': 'main',
        }
