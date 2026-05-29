# -*- coding: utf-8 -*-
"""Regression against client workbook ANACO_P1002_25_CON_REV03 (if present on disk)."""
import base64
import io
from pathlib import Path

from odoo.tests import tagged
from odoo.tests.common import TransactionCase

try:
    import openpyxl
except ImportError:  # pragma: no cover
    openpyxl = None

P1002_CANDIDATES = (
    Path(__file__).resolve().parents[2] / 'ANACO_P1002_25_CON_REV03_EIALL+ALL_P1.xlsx',
    Path(__file__).resolve().parents[2].parent / 'ANACO_P1002_25_CON_REV03_EIALL+ALL_P1.xlsx',
    Path(__file__).resolve().parents[2] / 'docs' / 'samples' / 'client'
    / 'ANACO_P1002_25_CON_REV03_EIALL+ALL_P1.xlsx',
)


@tagged('post_install', '-at_install')
class TestSbuP1002Workbook(TransactionCase):
    def _p1002_path(self):
        for path in P1002_CANDIDATES:
            if path.is_file():
                return path
        return None

    def test_p1002_import_row_counts(self):
        if not openpyxl:
            self.skipTest('openpyxl not installed')
        path = self._p1002_path()
        if not path:
            self.skipTest('P1002 workbook not on disk')
        partner = self.env['res.partner'].create({'name': 'P1002 import partner'})
        estimate = self.env['sbu.estimate'].create({'partner_id': partner.id})
        wizard = self.env['sbu.estimate.anaco.import.wizard'].create({
            'estimate_id': estimate.id,
            'data_file': base64.b64encode(path.read_bytes()),
            'data_filename': path.name,
            'import_anaco': True,
            'import_sal': True,
            'replace_anaco_lines': True,
            'replace_sal_lines': True,
            'auto_detect_first_row': True,
            'auto_detect_sal_first_row': True,
            'auto_detect_sal_columns': True,
        })
        wizard.action_import()
        self.assertGreaterEqual(len(estimate.line_ids), 30)
        self.assertLessEqual(len(estimate.line_ids), 40)
        self.assertGreaterEqual(len(estimate.sal_line_ids), 30)
        self.assertLessEqual(len(estimate.sal_line_ids), 40)
        self.assertFalse(
            estimate.line_ids.filtered(lambda l: 'ANALISI COSTO' in (l.description or '')),
        )
        self.assertFalse(
            estimate.sal_line_ids.filtered(
                lambda l: (l.description or '') == 'SERRAMENTI SERIE F' and not l.item_ref,
            ),
        )
        self.assertEqual(
            len(estimate.sal_line_ids.filtered(lambda l: (l.cumulative_pct or 0) > 0)),
            0,
        )

    def test_p1002_f1_excel_parity_bs_bb_bc(self):
        """Row F1: BS prezzo, BB/BC costo certificato Excel (tolleranza €0.05)."""
        if not openpyxl:
            self.skipTest('openpyxl not installed')
        path = self._p1002_path()
        if not path:
            self.skipTest('P1002 workbook not on disk')
        from odoo.addons.sbu_estimate.wizards.sbu_openpyxl_utils import load_openpyxl_workbook

        wb = load_openpyxl_workbook(path, data_only=True, read_only=True)
        anaco = wb['ANACO']
        excel_bs = float(anaco.cell(13, 71).value or 0)
        excel_bb = float(anaco.cell(13, 54).value or 0)
        excel_bc = float(anaco.cell(13, 55).value or 0)
        excel_qty = float(anaco.cell(13, 8).value or 1)
        wb.close()

        partner = self.env['res.partner'].create({'name': 'P1002 F1 parity'})
        estimate = self.env['sbu.estimate'].create({'partner_id': partner.id})
        wizard = self.env['sbu.estimate.anaco.import.wizard'].create({
            'estimate_id': estimate.id,
            'data_file': base64.b64encode(path.read_bytes()),
            'data_filename': path.name,
            'import_anaco': True,
            'import_sal': False,
            'replace_anaco_lines': True,
            'auto_detect_first_row': True,
        })
        wizard.action_import()
        line = estimate.line_ids.filtered(lambda l: (l.pos or '').strip() == 'F1')
        self.assertEqual(len(line), 1, line.mapped('pos'))
        line = line[0]
        self.assertAlmostEqual(line.price_anaco_bs_cad, excel_bs, places=2)
        self.assertAlmostEqual(line.price_total_cad, excel_bs, places=2)
        self.assertAlmostEqual(line.cost_anaco_bb_cad, excel_bb, places=2)
        self.assertAlmostEqual(line.cost_total_cad, excel_bb, places=2)
        self.assertAlmostEqual(line.cost_anaco_bc_tot, excel_bc, places=0)
        self.assertAlmostEqual(line.cost_total_tot, excel_bc, places=0)
        self.assertAlmostEqual(line.qty, excel_qty, places=2)
        expected_margin_pct = (
            (line.price_total_tot - line.cost_total_tot) / line.price_total_tot * 100.0
        )
        self.assertAlmostEqual(line.margin_pct, expected_margin_pct, places=1)
        self.assertAlmostEqual(line.cost_staffame_cad, 0.0, places=2)
