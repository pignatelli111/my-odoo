# -*- coding: utf-8 -*-
import base64
import io

from odoo.exceptions import UserError
from odoo.tests import tagged
from odoo.tests.common import TransactionCase

try:
    import openpyxl
except ImportError:  # pragma: no cover
    openpyxl = None

from odoo.addons.sbu_estimate.wizards.sbu_estimate_anaco_import_wizard import (
    ANACO_COL_COST_STAFFAME,
    _coerce_import_sal_pct,
    _detect_anaco_cost_staffame_column,
    _detect_sal_retention_column,
    _detect_sal_pct_start_column,
    _filter_model_fields,
    _sal_pct_plan_looks_incremental,
    SAL_COL_SAL_START,
)


@tagged('post_install', '-at_install')
class TestSbuAnacoImport(TransactionCase):
    def test_filter_model_fields_drops_unknown_keys(self):
        line_model = self.env['sbu.estimate.line']
        vals = {
            'description': 'Test',
            'qty': 1.0,
            'cost_nolo_cad': 10.0,
        }
        filtered = _filter_model_fields(line_model, vals)
        self.assertNotIn('cost_nolo_cad', filtered)
        self.assertEqual(filtered['description'], 'Test')

    def test_import_smoke_with_minimal_workbook(self):
        if not openpyxl:
            self.skipTest('openpyxl not installed')
        partner = self.env['res.partner'].create({'name': 'Import smoke partner'})
        estimate = self.env['sbu.estimate'].create({'partner_id': partner.id})
        wb = openpyxl.Workbook()
        default = wb.active
        wb.remove(default)
        sh = wb.create_sheet('ANACO')
        sh.cell(5, 11, 1.0)
        sh.cell(5, 12, 1.0)
        sh.cell(5, 13, 1.0)
        sh.cell(12, 2, 'FT01')
        sh.cell(12, 3, 'Finestra test')
        sh.cell(12, 8, 1)
        sh.cell(12, 71, 100.0)
        buf = io.BytesIO()
        wb.save(buf)
        wizard = self.env['sbu.estimate.anaco.import.wizard'].create({
            'estimate_id': estimate.id,
            'data_file': base64.b64encode(buf.getvalue()),
            'data_filename': 'smoke.xlsx',
            'import_anaco': True,
            'import_sal': False,
            'replace_anaco_lines': True,
        })
        wizard.action_import()
        self.assertEqual(len(estimate.line_ids), 1)
        self.assertEqual(estimate.line_ids[0].pos, 'FT01')

    def test_import_anaco_generates_distinta_bom(self):
        if not openpyxl:
            self.skipTest('openpyxl not installed')
        self.env['product.template'].create({
            'name': 'Coib test',
            'default_code': 'SBU-COIB',
            'type': 'consu',
            'purchase_ok': True,
        })
        partner = self.env['res.partner'].create({'name': 'BOM import partner'})
        estimate = self.env['sbu.estimate'].create({'partner_id': partner.id})
        wb = openpyxl.Workbook()
        default = wb.active
        wb.remove(default)
        sh = wb.create_sheet('ANACO')
        sh.cell(5, 11, 1.0)
        sh.cell(12, 2, 'FT01')
        sh.cell(12, 3, 'Finestra test')
        sh.cell(12, 8, 1)
        sh.cell(12, 30, 12.5)
        sh.cell(12, 71, 100.0)
        buf = io.BytesIO()
        wb.save(buf)
        wizard = self.env['sbu.estimate.anaco.import.wizard'].create({
            'estimate_id': estimate.id,
            'data_file': base64.b64encode(buf.getvalue()),
            'data_filename': 'bom.xlsx',
            'import_anaco': True,
            'import_sal': False,
            'import_distinta_bom': True,
            'replace_anaco_lines': True,
        })
        wizard.action_import()
        self.assertGreater(len(estimate.line_ids.bom_line_ids), 0)
        coib = estimate.line_ids.bom_line_ids.filtered(
            lambda b: b.product_id.default_code == 'SBU-COIB'
        )
        self.assertEqual(len(coib), 1)
        self.assertAlmostEqual(coib.unit_cost, 12.5)

    def test_coerce_import_sal_pct_rejects_amounts(self):
        self.assertIsNone(_coerce_import_sal_pct(44430.6))
        self.assertAlmostEqual(_coerce_import_sal_pct(25.0), 25.0)
        self.assertAlmostEqual(_coerce_import_sal_pct(0.25), 25.0)

    def test_sal_pct_plan_incremental_vs_single_100(self):
        self.assertFalse(_sal_pct_plan_looks_incremental([100.0]))
        self.assertTrue(_sal_pct_plan_looks_incremental([20.0, 30.0, 50.0]))

    def test_detect_sal_pct_start_column_non_rev7(self):
        if not openpyxl:
            self.skipTest('openpyxl not installed')
        wb = openpyxl.Workbook()
        default = wb.active
        wb.remove(default)
        sal = wb.create_sheet('Voci Contrattuali_SAL')
        sal.cell(10, 55, 'SAL-1')
        sal.cell(10, 56, 'SAL-2')
        sal.cell(10, 64, 'SAL-10')
        self.assertEqual(_detect_sal_pct_start_column(sal), 55)
        self.assertEqual(_detect_sal_pct_start_column(sal, None, SAL_COL_SAL_START), 55)

    def test_detect_sal_pct_does_not_treat_sal10_header_as_sal1(self):
        if not openpyxl:
            self.skipTest('openpyxl not installed')
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        sal = wb.create_sheet('Voci Contrattuali_SAL')
        sal.cell(8, 90, 'SAL-10')
        sal.cell(8, 91, 'SAL-11')
        col = _detect_sal_pct_start_column(sal, None, SAL_COL_SAL_START)
        self.assertEqual(col, SAL_COL_SAL_START)

    def test_import_sal_pct_from_detected_column(self):
        if not openpyxl:
            self.skipTest('openpyxl not installed')
        partner = self.env['res.partner'].create({'name': 'SAL col partner'})
        estimate = self.env['sbu.estimate'].create({'partner_id': partner.id})
        wb = openpyxl.Workbook()
        default = wb.active
        wb.remove(default)
        sal = wb.create_sheet('Voci Contrattuali_SAL')
        sal.cell(10, 55, 'SAL-1')
        sal.cell(10, 56, 'SAL-2')
        sal.cell(16, 2, 'SAL-TEST')
        sal.cell(16, 3, 'Voce test SAL %')
        sal.cell(16, 4, 1)
        sal.cell(16, 8, 10000.0)
        sal.cell(16, 55, 20.0)
        sal.cell(16, 56, 30.0)
        buf = io.BytesIO()
        wb.save(buf)
        wizard = self.env['sbu.estimate.anaco.import.wizard'].create({
            'estimate_id': estimate.id,
            'data_file': base64.b64encode(buf.getvalue()),
            'data_filename': 'sal_col.xlsx',
            'import_anaco': False,
            'import_sal': True,
            'replace_sal_lines': True,
            'auto_detect_sal_columns': True,
            'sal_first_row': 16,
        })
        wizard.action_import()
        self.assertEqual(len(estimate.sal_line_ids), 1)
        line = estimate.sal_line_ids[0]
        self.assertAlmostEqual(line.sal_1_pct, 20.0)
        self.assertAlmostEqual(line.sal_2_pct, 30.0)
        self.assertAlmostEqual(line.cumulative_pct, 50.0)

    def test_detect_staffame_column_from_header(self):
        if not openpyxl:
            self.skipTest('openpyxl not installed')
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        sh = wb.create_sheet('ANACO')
        sh.cell(8, 55, 'ST/LZ Staffame')
        self.assertEqual(_detect_anaco_cost_staffame_column(sh), 55)

    def test_import_staffame_cost_and_bs_price_parity(self):
        """Cosimo punto 12: staffame in cost total; BS drives customer price."""
        if not openpyxl:
            self.skipTest('openpyxl not installed')
        partner = self.env['res.partner'].create({'name': 'Parity import partner'})
        estimate = self.env['sbu.estimate'].create({'partner_id': partner.id})
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        sh = wb.create_sheet('ANACO')
        sh.cell(5, 11, 1.0)
        sh.cell(5, 12, 1.0)
        sh.cell(5, 13, 1.0)
        sh.cell(12, 2, 'F1')
        sh.cell(12, 3, 'Facade line')
        sh.cell(12, 4, 1000)
        sh.cell(12, 6, 1000)
        sh.cell(12, 8, 2)
        sh.cell(12, 30, 100.0)
        sh.cell(12, 16, 25.0)
        sh.cell(12, 71, 500.0)
        buf = io.BytesIO()
        wb.save(buf)
        wizard = self.env['sbu.estimate.anaco.import.wizard'].create({
            'estimate_id': estimate.id,
            'data_file': base64.b64encode(buf.getvalue()),
            'data_filename': 'parity.xlsx',
            'import_anaco': True,
            'import_sal': False,
            'replace_anaco_lines': True,
        })
        wizard.action_import()
        line = estimate.line_ids
        self.assertEqual(len(line), 1)
        self.assertAlmostEqual(line.cost_staffame_cad, 25.0)
        self.assertAlmostEqual(line.price_anaco_bs_cad, 500.0)
        self.assertAlmostEqual(line.price_total_cad, 500.0)
        self.assertAlmostEqual(line.cost_total_cad, 125.0)
        self.assertAlmostEqual(line.margin_pct, 75.0)

    def test_import_sal_retention_from_detected_column(self):
        if not openpyxl:
            self.skipTest('openpyxl not installed')
        partner = self.env['res.partner'].create({'name': 'Retention import partner'})
        estimate = self.env['sbu.estimate'].create({'partner_id': partner.id})
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        sal = wb.create_sheet('Voci Contrattuali_SAL')
        sal.cell(10, 11, 'Ritenuta %')
        sal.cell(16, 2, 'S1')
        sal.cell(16, 3, 'Voce con garanzia')
        sal.cell(16, 4, 1)
        sal.cell(16, 8, 10000.0)
        sal.cell(16, 11, 8.0)
        buf = io.BytesIO()
        wb.save(buf)
        wizard = self.env['sbu.estimate.anaco.import.wizard'].create({
            'estimate_id': estimate.id,
            'data_file': base64.b64encode(buf.getvalue()),
            'data_filename': 'ret.xlsx',
            'import_anaco': False,
            'import_sal': True,
            'replace_sal_lines': True,
            'import_sal_retention': True,
            'auto_detect_sal_columns': True,
        })
        wizard.action_import()
        self.assertEqual(len(estimate.sal_line_ids), 1)
        self.assertEqual(_detect_sal_retention_column(sal), 11)
        self.assertAlmostEqual(estimate.sal_line_ids.retention_percent, 8.0)

    def test_import_requires_file(self):
        partner = self.env['res.partner'].create({'name': 'Import no-file partner'})
        estimate = self.env['sbu.estimate'].create({'partner_id': partner.id})
        wizard = self.env['sbu.estimate.anaco.import.wizard'].create({
            'estimate_id': estimate.id,
            'import_anaco': True,
            'import_sal': False,
        })
        with self.assertRaises(UserError):
            wizard.action_import()
