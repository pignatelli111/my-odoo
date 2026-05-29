#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Compare P1002 Excel ANACO row totals vs expected Odoo import targets (BS, BB, BC).

Usage:
  python tools/p1002_parity_check.py "f:\\TASK\\20 . Odoo\\ANACO_P1002_25_CON_REV03_EIALL+ALL_P1.xlsx"
"""
from __future__ import annotations

import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    raise SystemExit('pip install openpyxl') from None

COL_POS, COL_QTY, COL_BS, COL_BB, COL_BC = 2, 8, 71, 54, 55


def main():
    path = Path(sys.argv[1] if len(sys.argv) > 1 else r'ANACO_P1002_25_CON_REV03_EIALL+ALL_P1.xlsx')
    if not path.is_file():
        print(f'File not found: {path}')
        return 1
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sh = wb['ANACO']
    print(f'File: {path.name}\n')
    print(f'{"Pos":<8} {"BS":>12} {"BB":>12} {"BC":>14} {"Margin%":>8}')
    n = 0
    for r in range(13, (sh.max_row or 13) + 1):
        pos = sh.cell(r, COL_POS).value
        if not pos or str(pos).strip() in ('ITEMS', 'COD.', 'RIFERIMENTO', 'OGGETTO'):
            continue
        pos_s = str(pos).split('\n')[0].strip()
        if not pos_s or pos_s[0].isdigit() is False and pos_s[0].upper() not in 'FLA':
            if 'SERRAMENTI' in str(sh.cell(r, 3).value or ''):
                continue
        bs = float(sh.cell(r, COL_BS).value or 0)
        bb = float(sh.cell(r, COL_BB).value or 0)
        bc = float(sh.cell(r, COL_BC).value or 0)
        if bs <= 0 and bb <= 0:
            continue
        margin = (bs * (sh.cell(r, COL_QTY).value or 1) - bc) / (bs * (sh.cell(r, COL_QTY).value or 1)) * 100 if bc else 0
        print(f'{pos_s:<8} {bs:12.2f} {bb:12.2f} {bc:14.2f} {margin:7.1f}%')
        n += 1
        if n >= 40:
            break
    wb.close()
    print(f'\n{n} product rows with BS/BB shown.')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
