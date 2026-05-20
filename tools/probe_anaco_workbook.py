#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Dry-run analysis of a client ANACO .xlsx (same layout rules as the Odoo import wizard).

Usage:
  cd my-odoo
  pip install openpyxl
  python tools/probe_anaco_workbook.py "docs/samples/client/ANACO_P1002_25_CON_REV03_EIALL+ALL_P1.xlsx"

Copy the client file into docs/samples/client/ first (not committed by default).
"""
from __future__ import annotations

import math
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    raise SystemExit("Install openpyxl: pip install openpyxl") from None

# Mirror sbu_estimate_anaco_import_wizard.py (1-based columns)
ANACO_ROW_PARAMS = 5
ANACO_ROW_FIRST_DATA_DEFAULT = 12
ANACO_COL_POS = 2
ANACO_COL_DESC = 3
ANACO_COL_B_MM = 4
ANACO_COL_H_MM = 6
ANACO_COL_QTY = 8
ANACO_COL_SC_MULT = (11, 12, 13)
ANACO_COL_BS_UNIT = 71
ANACO_COL_COST_IND_PCT = 65
ANACO_COL_COST_MOL_PCT = 68
ANACO_COL_PRICE = {
    'price_serramento_cad': 14,
    'price_accessori_cad': 22,
    'price_vetro_cad': 40,
}
ANACO_COL_COST = {
    'cost_coibentazione_cad': 30,
    'cost_trasporto_cad': 57,
}

SAL_ROW_FIRST_DATA_DEFAULT = 16
SAL_COL_ITEM = 2
SAL_COL_DESC = 3
SAL_COL_QTY = 4
SAL_COL_UNIT_PRICE = 8
SAL_COL_SAL_START = 98  # REV7 default; client files often differ


def _sal_header_normalize(text):
    import re
    raw = (text or '').strip().upper()
    compact = re.sub(r'[\s._\-/]+', '', raw)
    return compact, raw


def _sal_header_is_sal_n(compact, n):
    import re
    return bool(re.fullmatch(rf'SAL0?{n}', compact))


def _detect_sal_pct_start_column(sh, fallback=SAL_COL_SAL_START):
    import re
    max_col = min((sh.max_column or 120) + 1, 160)
    for row in range(1, 30):
        for col in range(1, max_col):
            compact, raw = _sal_header_normalize(_cell_str(sh, row, col))
            is_sal1 = _sal_header_is_sal_n(compact, 1) or bool(
                re.search(r'SAL\s*[-.]?\s*1(?!\d)', raw)
            )
            if not is_sal1:
                continue
            for off in range(1, 5):
                c2, _ = _sal_header_normalize(_cell_str(sh, row, col + off))
                if _sal_header_is_sal_n(c2, 2):
                    return col
            if _sal_header_is_sal_n(compact, 1):
                return col
    return fallback

OFFERTA_ROW_FIRST_DATA_DEFAULT = 24
OFFERTA_COL_CODE = 2
OFFERTA_COL_DESC = 5
OFFERTA_COL_QTY = 7
OFFERTA_COL_PRICE = 8

_ANACO_SKIP_B = frozenset({
    'POS.', 'COD.', 'ITEMS', 'DATA', 'RIFERIMENTO', 'OGGETTO', 'CLIENTE',
})
_ANACO_NUMERIC_COLS = tuple(
    set(ANACO_COL_PRICE.values())
    | {ANACO_COL_B_MM, ANACO_COL_H_MM, ANACO_COL_QTY, ANACO_COL_BS_UNIT}
    | set(ANACO_COL_COST.values())
)


def _cell_str(sh, row, col):
    v = sh.cell(row, col).value
    if v is None:
        return ''
    if isinstance(v, (int, float)):
        if v == 0:
            return ''
        if float(v).is_integer():
            return str(int(v))
        return str(v).strip()
    return str(v).strip()


def _cell_num(sh, row, col):
    v = sh.cell(row, col).value
    if v is None or v == '':
        return None
    if isinstance(v, str):
        vs = v.replace(',', '.').strip()
        if not vs or vs == '-' or vs.upper().startswith('#'):
            return None
        v = vs
    try:
        x = float(v)
    except (TypeError, ValueError):
        return None
    if isinstance(x, float) and (math.isnan(x) or math.isinf(x)):
        return None
    return x


def _sheet_by_aliases(wb, aliases):
    lower = {n.lower().replace(' ', '_'): n for n in wb.sheetnames}
    for a in aliases:
        key = a.lower().replace(' ', '_')
        if key in lower:
            return wb[lower[key]]
    return None


def _anaco_row_is_data(sh, row):
    pos = _cell_str(sh, row, ANACO_COL_POS).upper()
    if pos and pos not in _ANACO_SKIP_B and not pos.startswith('CLIENTE'):
        return True
    desc = _cell_str(sh, row, ANACO_COL_DESC)
    if desc and desc.upper() not in ("DESCRIZIONE D'OFFERTA", 'DESCRIZIONE'):
        return True
    b_mm = _cell_num(sh, row, ANACO_COL_B_MM)
    h_mm = _cell_num(sh, row, ANACO_COL_H_MM)
    if (b_mm and b_mm > 0) or (h_mm and h_mm > 0):
        return True
    qty = _cell_num(sh, row, ANACO_COL_QTY)
    if qty and qty > 0:
        return True
    bs = _cell_num(sh, row, ANACO_COL_BS_UNIT)
    if bs and bs > 0:
        return True
    for col in _ANACO_NUMERIC_COLS:
        v = _cell_num(sh, row, col)
        if v and v > 0:
            return True
    return False


def _detect_anaco_first_row(sh, fallback):
    for row in range(1, 81):
        if _anaco_row_is_data(sh, row):
            return row
    return fallback


def _count_anaco_rows(sh, first_row):
    count = 0
    empty_run = 0
    max_row = sh.max_row or first_row
    samples = []
    for r in range(first_row, max_row + 1):
        if not _anaco_row_is_data(sh, r):
            empty_run += 1
            if empty_run >= 8:
                break
            continue
        empty_run = 0
        count += 1
        if len(samples) < 5:
            samples.append(
                (r, _cell_str(sh, r, ANACO_COL_POS), _cell_str(sh, r, ANACO_COL_DESC)[:60])
            )
    return count, samples


def _count_sal_rows(sh, first_row, sal_pct_col=SAL_COL_SAL_START):
    count = 0
    empty_run = 0
    samples = []
    max_row = sh.max_row or first_row
    for r in range(first_row, max_row + 1):
        item = _cell_str(sh, r, SAL_COL_ITEM)
        desc = _cell_str(sh, r, SAL_COL_DESC)
        if not item and not desc:
            empty_run += 1
            if empty_run >= 5:
                break
            continue
        empty_run = 0
        count += 1
        if len(samples) < 5:
            qty = _cell_num(sh, r, SAL_COL_QTY)
            unit = _cell_num(sh, r, SAL_COL_UNIT_PRICE)
            sal1 = _cell_num(sh, r, sal_pct_col)
            samples.append((r, item, desc[:50], qty, unit, sal1))
    return count, samples


def _count_offerta_rows(sh):
    first_row = OFFERTA_ROW_FIRST_DATA_DEFAULT
    for row in range(20, 41):
        if _cell_str(sh, row, OFFERTA_COL_CODE).upper() == 'CODICE':
            first_row = row + 1
            break
    count = 0
    empty_run = 0
    max_row = sh.max_row or first_row
    for r in range(first_row, max_row + 1):
        code = _cell_str(sh, r, OFFERTA_COL_CODE)
        desc = _cell_str(sh, r, OFFERTA_COL_DESC)
        qty = _cell_num(sh, r, OFFERTA_COL_QTY)
        price = _cell_num(sh, r, OFFERTA_COL_PRICE)
        if code.upper() in ('CODICE', 'ITEM', '0'):
            continue
        if not code and not desc and not (qty and qty > 0) and not (price and price > 0):
            empty_run += 1
            if empty_run >= 8:
                break
            continue
        empty_run = 0
        count += 1
    return count, first_row


def probe(path: Path) -> int:
    if not path.is_file():
        print(f"ERROR: file not found: {path}")
        print("Copy the client workbook to e.g. docs/samples/client/ and re-run.")
        return 1

    print(f"=== File: {path.name} ({path.stat().st_size // 1024} KB) ===\n")

    wb_vals = openpyxl.load_workbook(path, data_only=True, read_only=True)
    wb_form = openpyxl.load_workbook(path, data_only=False, read_only=True)

    print("Sheets:", ", ".join(wb_vals.sheetnames))

    anaco = _sheet_by_aliases(wb_vals, ('ANACO', 'Anaco'))
    anaco_form = _sheet_by_aliases(wb_form, ('ANACO', 'Anaco'))
    sal = _sheet_by_aliases(wb_vals, (
        'Voci Contrattuali_SAL',
        'Voci Contrattuali SAL',
        'VOCI CONTRATTUALI_SAL',
    ))
    offerta = _sheet_by_aliases(wb_vals, ('OFFERTA', 'Offerta'))

    if not anaco:
        print("\nERROR: no ANACO sheet — wizard will fail ANACO import.")
        wb_vals.close()
        wb_form.close()
        return 1

    detected = _detect_anaco_first_row(anaco, ANACO_ROW_FIRST_DATA_DEFAULT)
    n_anaco_vals, samples = _count_anaco_rows(anaco, detected)
    n_anaco_form, _ = _count_anaco_rows(anaco_form, detected) if anaco_form else (0, [])

    sc = [_cell_num(anaco, ANACO_ROW_PARAMS, c) for c in ANACO_COL_SC_MULT]
    ind = _cell_num(anaco, ANACO_ROW_PARAMS, ANACO_COL_COST_IND_PCT)
    mol = _cell_num(anaco, ANACO_ROW_PARAMS, ANACO_COL_COST_MOL_PCT)

    print(f"\n--- ANACO (values, data_only=True) ---")
    print(f"  First data row (auto-detect): {detected} (default wizard: {ANACO_ROW_FIRST_DATA_DEFAULT})")
    print(f"  Importable rows (estimate):   {n_anaco_vals}")
    if n_anaco_form != n_anaco_vals:
        print(f"  Rows with formulas only:     {n_anaco_form} (values) vs {n_anaco_vals} — open in Excel & save if 0 values")
    print(f"  Row 5 K/L/M multipliers:     {sc}")
    print(f"  Row 5 BM industrial %:       {ind}")
    print(f"  Row 5 BP MOL %:              {mol}")
    if samples:
        print("  Sample rows (excel row, pos, desc):")
        for row, pos, desc in samples:
            print(f"    {row}: {pos!r} | {desc!r}")

    if n_anaco_vals == 0 and offerta:
        n_off, off_first = _count_offerta_rows(offerta)
        print(f"\n--- OFFERTA fallback (if ANACO import finds 0 rows) ---")
        print(f"  First data row: {off_first}")
        print(f"  Importable rows: {n_off}")

    if sal:
        sal_pct_col = _detect_sal_pct_start_column(sal, SAL_COL_SAL_START)
        n_sal, sal_samples = _count_sal_rows(
            sal, SAL_ROW_FIRST_DATA_DEFAULT, sal_pct_col=sal_pct_col,
        )
        print(f"\n--- Voci Contrattuali_SAL ---")
        print(f"  First data row (wizard default): {SAL_ROW_FIRST_DATA_DEFAULT}")
        print(f"  SAL-1 column (detected): {sal_pct_col} (REV7 default: {SAL_COL_SAL_START})")
        if sal_pct_col == SAL_COL_SAL_START and not any(s[-1] for s in sal_samples):
            print("  WARN: SAL-1 header not found / no % in col 98 — Cum.% may stay 0 in Odoo.")
        print(f"  Importable SAL lines: {n_sal}")
        if sal_samples:
            print("  Sample (row, item, desc, qty, unit, SAL-1% at detected col):")
            for s in sal_samples:
                print(f"    {s}")
    else:
        print("\nWARN: no SAL sheet — disable «Importa Voci Contrattuali SAL» or fix sheet name.")

    print("\n--- Odoo import checklist ---")
    print("  1. Nuovo preventivo -> Salva")
    print("  2. Importa Excel ANACO -> upload this file")
    print("  3. Importa righe ANACO + SAL, Sostituisci entrambe, Rileva prima riga, Fallback OFFERTA")
    if detected != ANACO_ROW_FIRST_DATA_DEFAULT:
        print(f"  4. If needed set «Prima riga dati ANACO» = {detected}")
    print(f"  5. Prima riga SAL = {SAL_ROW_FIRST_DATA_DEFAULT} (change if probe shows data elsewhere)")

    wb_vals.close()
    wb_form.close()
    return 0 if n_anaco_vals or (offerta and _count_offerta_rows(offerta)[0]) else 1


def main():
    if len(sys.argv) < 2:
        default = (
            Path(__file__).resolve().parents[1]
            / "docs"
            / "samples"
            / "client"
            / "ANACO_P1002_25_CON_REV03_EIALL+ALL_P1.xlsx"
        )
        print(f"No path given; trying default:\n  {default}\n")
        path = default
    else:
        path = Path(sys.argv[1])
    raise SystemExit(probe(path))


if __name__ == "__main__":
    main()
