"""Dump structure of SBU-related Excel files (run: python tools/dump_excel_workbooks.py)."""
from __future__ import annotations

import json
from pathlib import Path

import openpyxl

ROOT = Path(r"f:\TASK\20 . Odoo\New folder")
FILES = [
    ROOT / "Definizioni generali per ERP.xlsx",
    ROOT / "CARTELLA PREVENTIVO" / "ANALISI E PREVENTIVO" / "ANACO_REV7_111122.xlsx",
    ROOT / "2_2_0_COMPLIANCE COMMESSA" / "2_PM_BDG OP_GANTT_SAL_SOTTOMISSIONI" / "2_3_SOTTOMISSIONI" / "STD_SBB_SCHEDASOTTOMISSIONE.xlsx",
    ROOT / "2_2_0_COMPLIANCE COMMESSA" / "4_TECH" / "4_10_ORDINI INTERNI" / "4_10_1_RDA" / "P0105_20_RDA-044_00.xlsx",
    ROOT / "2_2_0_COMPLIANCE COMMESSA" / "4_TECH" / "4_10_ORDINI INTERNI" / "4_10_2_ACP" / "P0105_2020 ACP 08_00.xlsx",
    ROOT / "2_2_0_COMPLIANCE COMMESSA" / "4_TECH" / "4_10_ORDINI INTERNI" / "4_10_3_ACO" / "P0105_20_ACO-001_01.xlsx",
    ROOT / "2_2_0_COMPLIANCE COMMESSA" / "4_TECH" / "4_10_ORDINI INTERNI" / "4_10_4_LDS" / "P105-2020-LDS06_00.xlsx",
]


def cell_preview(v, max_len=72):
    if v is None:
        return ""
    if isinstance(v, str) and v.startswith("="):
        return (v[:max_len] + "...") if len(v) > max_len else v
    s = str(v).replace("\n", " | ")
    return (s[:max_len] + "...") if len(s) > max_len else s


def scan_sheet(ws, max_row=18, max_col=22):
    rows = []
    lim_r = min(max_row, ws.max_row or 1)
    lim_c = min(max_col, ws.max_column or 1)
    for r in range(1, lim_r + 1):
        row = []
        for c in range(1, lim_c + 1):
            row.append(cell_preview(ws.cell(row=r, column=c).value))
        if any(row):
            rows.append({"r": r, "cells": row})
    return rows


def main():
    out = []
    for path in FILES:
        rec = {"path": str(path), "exists": path.is_file(), "sheets": [], "error": None}
        if not path.is_file():
            out.append(rec)
            continue
        try:
            wb = openpyxl.load_workbook(path, data_only=False, read_only=True)
            rec["sheets"] = wb.sheetnames
            for i, name in enumerate(wb.sheetnames[:6]):
                ws = wb[name]
                safe = "".join(c if c.isalnum() or c in " -_" else "_" for c in name)[:40]
                rec[f"sheet_{i}_{safe}"] = {
                    "max_row": ws.max_row,
                    "max_column": ws.max_column,
                    "preview": scan_sheet(ws),
                }
            wb.close()
        except Exception as e:
            rec["error"] = repr(e)
        out.append(rec)

    dest = Path(__file__).resolve().parents[1] / "docs" / "excel_workbooks_inventory.json"
    dest.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {dest}")


if __name__ == "__main__":
    main()
