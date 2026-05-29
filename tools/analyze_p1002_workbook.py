#!/usr/bin/env python3
"""Analyze P1002 ANACO workbook vs sbu_estimate import wizard layout."""
import re
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    raise SystemExit('pip install openpyxl') from None

# Mirror wizard constants
ANACO_ROW_PARAMS = 5
ANACO_COL_POS, ANACO_COL_DESC = 2, 3
ANACO_COL_B_MM, ANACO_COL_H_MM, ANACO_COL_QTY = 4, 6, 8
ANACO_COL_BS_UNIT = 71
SAL_ROW_FIRST = 16
SAL_COL_ITEM, SAL_COL_DESC, SAL_COL_QTY, SAL_COL_UNIT = 2, 3, 4, 8
SAL_COL_SAL_START = 98


def cell_str(v):
    if v is None:
        return ''
    if isinstance(v, (int, float)):
        if v == 0:
            return ''
        return str(int(v)) if float(v).is_integer() else str(v)
    return str(v).strip()


def sal_header_key(text):
    raw = cell_str(text).upper()
    compact = re.sub(r'[\s._\-/]+', '', raw)
    return compact, raw


def find_sal_headers(sh, max_row=30, max_col=120):
    hits = []
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            compact, raw = sal_header_key(sh.cell(row, col).value)
            if re.fullmatch(r'SAL0?(\d+)', compact) or re.search(r'SAL\s*[-.]?\s*(\d+)(?!\d)', raw):
                m = re.fullmatch(r'SAL0?(\d+)', compact)
                n = int(m.group(1)) if m else None
                if n is None:
                    m2 = re.search(r'SAL\s*[-.]?\s*(\d+)(?!\d)', raw)
                    n = int(m2.group(1)) if m2 else None
                hits.append((row, col, get_column_letter(col), raw or compact, n))
    return hits


def row_sample(sh, row, cols):
    return {c: sh.cell(row, c).value for c in cols}


def main():
    path = Path(sys.argv[1] if len(sys.argv) > 1 else r'F:\TASK\20 . Odoo\ANACO_P1002_25_CON_REV03_EIALL+ALL_P1.xlsx')
    print(f'File: {path} ({path.stat().st_size // 1024} KB)\n')
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    wb_f = openpyxl.load_workbook(path, data_only=False, read_only=True)
    print('Sheets:', ', '.join(wb.sheetnames))

    # ANACO
    anaco = None
    for name in wb.sheetnames:
        if name.upper() == 'ANACO':
            anaco = wb[name]
            break
    if anaco:
        print('\n=== ANACO ===')
        print(f'max_row={anaco.max_row} max_col={anaco.max_column}')
        print(f'Row {ANACO_ROW_PARAMS} K/L/M:', [anaco.cell(ANACO_ROW_PARAMS, c).value for c in (11, 12, 13)])
        print(f'Row {ANACO_ROW_PARAMS} BM/BP (65/68):', anaco.cell(ANACO_ROW_PARAMS, 65).value, anaco.cell(ANACO_ROW_PARAMS, 68).value)
        for r in range(8, 20):
            pos = cell_str(anaco.cell(r, ANACO_COL_POS).value)
            desc = cell_str(anaco.cell(r, ANACO_COL_DESC).value)[:50]
            qty = anaco.cell(r, ANACO_COL_QTY).value
            bs = anaco.cell(r, ANACO_COL_BS_UNIT).value
            if pos or desc or qty or bs:
                print(f'  r{r}: B={pos!r} C={desc!r} H_qty={qty} BS71={bs}')

    # SAL
    sal = None
    for name in wb.sheetnames:
        if 'contrattuali' in name.lower() and 'sal' in name.lower():
            sal = wb[name]
            sal_name = name
            break
    if sal:
        print(f'\n=== {sal_name} ===')
        print(f'max_row={sal.max_row} max_col={sal.max_column}')
        headers = find_sal_headers(sal, max_row=25, max_col=min(sal.max_column or 120, 120))
        sal1 = [h for h in headers if h[4] == 1]
        print(f'SAL-1 header hits ({len(sal1)}):')
        for h in sal1[:8]:
            print(f'  row {h[0]} col {h[1]} ({h[2]}): {h[3]!r}')
        # consecutive SAL-1..10 on same row?
        by_row = {}
        for row, col, letter, raw, n in headers:
            if n and 1 <= n <= 10:
                by_row.setdefault(row, {})[n] = (col, letter, raw)
        for row in sorted(by_row):
            nums = sorted(by_row[row])
            if len(nums) >= 3:
                cols = [by_row[row][n][0] for n in nums]
                contiguous = cols == list(range(min(cols), min(cols) + len(cols)))
                print(f'  Header row {row}: SAL-{nums[0]}..{nums[-1]} cols {[by_row[row][n][1] for n in nums]} contiguous={contiguous}')
        # sample data rows 14-20
        sal1_col = sal1[0][1] if sal1 else SAL_COL_SAL_START
        print(f'\nData sample (wizard defaults: first_row={SAL_ROW_FIRST}, sal_start={SAL_COL_SAL_START}, detected SAL-1 col={sal1_col}):')
        for r in range(14, 22):
            item = cell_str(sal.cell(r, SAL_COL_ITEM).value)
            desc = cell_str(sal.cell(r, SAL_COL_DESC).value)[:40]
            if not item and not desc:
                continue
            pcts_wiz = [sal.cell(r, SAL_COL_SAL_START + i).value for i in range(10)]
            pcts_det = [sal.cell(r, sal1_col + i).value for i in range(10)] if sal1_col else []
            print(f'  r{r}: item={item!r} desc={desc!r}')
            print(f'       unit@8={sal.cell(r, 8).value} qty@4={sal.cell(r, 4).value}')
            print(f'       pct@98-107 (wiz): {pcts_wiz[:5]}...')
            if sal1_col != SAL_COL_SAL_START:
                print(f'       pct@{sal1_col}+ (det): {pcts_det[:5]}...')
        # scan wide for % block on row 16
        if sal.max_row >= 16:
            r = 16
            blocks = []
            for start in range(48, min((sal.max_column or 110) + 1, 110)):
                vals = [sal.cell(r, start + i).value for i in range(10)]
                nums = [v for v in vals if isinstance(v, (int, float)) and v is not None]
                if len(nums) >= 2:
                    s = sum(float(x) if abs(float(x)) <= 1 else float(x) for x in nums)
                    if 0 < s <= 150:
                        blocks.append((start, get_column_letter(start), vals, s))
            print(f'\n  Row {r} numeric 10-col blocks (48-110) with 2+ nums:')
            for b in blocks[:6]:
                print(f'    start {b[0]} ({b[1]}): {b[2]} sum~{b[3]:.1f}')

    # OFFERTA
    for name in wb.sheetnames:
        if name.upper() == 'OFFERTA':
            off = wb[name]
            print(f'\n=== OFFERTA === max_row={off.max_row}')
            for r in range(20, 28):
                if cell_str(off.cell(r, 2).value).upper() == 'CODICE':
                    print(f'  CODICE header row {r}')
            break

    wb.close()
    wb_f.close()


if __name__ == '__main__':
    main()
