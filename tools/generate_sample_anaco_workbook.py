#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generate a rich ANACO-style .xlsx for SBU import wizard testing.

Layout matches sbu_estimate/wizards/sbu_estimate_anaco_import_wizard.py (REV7 columns).

Usage:
  cd my-odoo
  pip install openpyxl   # if needed
  python tools/generate_sample_anaco_workbook.py

Output:
  docs/samples/SBU_ANACO_SAMPLE_UAT.xlsx
"""
from pathlib import Path

try:
    from openpyxl import Workbook
except ImportError:
    raise SystemExit("Install openpyxl: pip install openpyxl") from None

# Same 1-based indices as the wizard
ANACO_ROW_PARAMS = 5
ANACO_COL_POS = 2
ANACO_COL_DESC = 3
ANACO_COL_B_MM = 4
ANACO_COL_H_MM = 6
ANACO_COL_QTY = 8
ANACO_COL_SC_MULT = (11, 12, 13)
ANACO_COL_BS_UNIT = 71
ANACO_COL_PRICE_SERR = 14
ANACO_COL_PRICE_ACCESSORI = 22
ANACO_COL_PRICE_VETRO = 40
ANACO_COL_COST_COIB = 30
ANACO_COL_COST_POSA_LIN = 50
ANACO_COL_COST_TRASPORTO = 57
ANACO_COL_COST_TECH_PM = 59
ANACO_COL_COST_CANTIERE = 61
ANACO_COL_COST_EXTRA = 63
ANACO_COL_COST_IND_PCT = 65
ANACO_COL_COST_MOL_PCT = 68
ANACO_COL_NOTE = 74

SAL_ROW_FIRST = 16
SAL_COL_ITEM = 2
SAL_COL_DESC = 3
SAL_COL_QTY = 4
SAL_COL_UNIT_PRICE = 8
SAL_COL_SAL_START = 98

OFFERTA_COL_CODE = 2
OFFERTA_COL_WIDTH = 3
OFFERTA_COL_HEIGHT = 4
OFFERTA_COL_DESC = 5
OFFERTA_COL_UOM = 6
OFFERTA_COL_QTY = 7
OFFERTA_COL_PRICE = 8


def _anaco_line(ws, row, pos, desc, b_mm, h_mm, qty, extra=None):
    extra = extra or {}
    ws.cell(row, ANACO_COL_POS, pos)
    ws.cell(row, ANACO_COL_DESC, desc)
    if b_mm is not None:
        ws.cell(row, ANACO_COL_B_MM, b_mm)
    if h_mm is not None:
        ws.cell(row, ANACO_COL_H_MM, h_mm)
    ws.cell(row, ANACO_COL_QTY, qty)
    for col, val in extra.items():
        if val is not None:
            ws.cell(row, col, val)


def main():
    root = Path(__file__).resolve().parents[1]
    out_dir = root / "docs" / "samples"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / "SBU_ANACO_SAMPLE_UAT.xlsx"

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    anaco = wb.create_sheet("ANACO")

    # Row 5: global multipliers (K/L/M) -> successive discounts in Odoo
    anaco.cell(ANACO_ROW_PARAMS, ANACO_COL_SC_MULT[0], 0.95)
    anaco.cell(ANACO_ROW_PARAMS, ANACO_COL_SC_MULT[1], 0.96)
    anaco.cell(ANACO_ROW_PARAMS, ANACO_COL_SC_MULT[2], 0.97)
    # Industrial % (BM) and MOL % (BP) on material subtotal in Odoo
    anaco.cell(ANACO_ROW_PARAMS, ANACO_COL_COST_IND_PCT, 6.0)
    anaco.cell(ANACO_ROW_PARAMS, ANACO_COL_COST_MOL_PCT, 4.0)

    # Header row (skipped by wizard)
    anaco.cell(11, ANACO_COL_POS, "COD.")
    anaco.cell(11, ANACO_COL_DESC, "DESCRIZIONE")

    # Line 12 — FT window: rich CAD price split + cost stack + BS unit price
    _anaco_line(
        anaco,
        12,
        "FT-SAMPLE-01",
        "Finestra campione — UAT (prezzi + costi)",
        1200,
        1500,
        2,
        {
            ANACO_COL_PRICE_SERR: 195.0,
            ANACO_COL_PRICE_ACCESSORI: 48.5,
            ANACO_COL_PRICE_VETRO: 210.0,
            ANACO_COL_COST_COIB: 42.0,
            ANACO_COL_COST_POSA_LIN: 88.0,
            ANACO_COL_COST_TRASPORTO: 36.0,
            ANACO_COL_COST_TECH_PM: 125.0,
            ANACO_COL_COST_CANTIERE: 55.0,
            ANACO_COL_COST_EXTRA: 22.0,
            ANACO_COL_BS_UNIT: 450.0,
            ANACO_COL_NOTE: "Campione UAT — nota riga 12",
        },
    )

    # Line 13 — LA skylight
    _anaco_line(
        anaco,
        13,
        "LA-SAMPLE-02",
        "Lucernario campione — UAT",
        1000,
        1000,
        1,
        {
            ANACO_COL_PRICE_SERR: 118.0,
            ANACO_COL_PRICE_VETRO: 95.0,
            ANACO_COL_COST_COIB: 28.0,
            ANACO_COL_COST_POSA_LIN: 52.0,
            ANACO_COL_COST_TRASPORTO: 18.0,
            ANACO_COL_COST_TECH_PM: 40.0,
            ANACO_COL_COST_CANTIERE: 30.0,
            ANACO_COL_COST_EXTRA: 10.0,
            ANACO_COL_BS_UNIT: 890.0,
        },
    )

    # Line 14 — accessories / Nr (no BxH): tests nr + cost_family from ACC prefix
    _anaco_line(
        anaco,
        14,
        "ACC-SAMPLE-03",
        "Kit ferramenta e guarnizioni — UAT (a pezzo)",
        None,
        None,
        5,
        {
            ANACO_COL_PRICE_ACCESSORI: 12.4,
            ANACO_COL_COST_EXTRA: 8.5,
            ANACO_COL_BS_UNIT: 120.0,
        },
    )

    sal = wb.create_sheet("Voci Contrattuali_SAL")
    # Contract line 1 — full 100% plan
    sal.cell(SAL_ROW_FIRST, SAL_COL_ITEM, "SAL-VOCE-01")
    sal.cell(SAL_ROW_FIRST, SAL_COL_DESC, "Avanzamento lavori — corpo principale")
    sal.cell(SAL_ROW_FIRST, SAL_COL_QTY, 1)
    sal.cell(SAL_ROW_FIRST, SAL_COL_UNIT_PRICE, 50000.0)
    sal.cell(SAL_ROW_FIRST, SAL_COL_SAL_START, 20.0)
    sal.cell(SAL_ROW_FIRST, SAL_COL_SAL_START + 1, 30.0)
    sal.cell(SAL_ROW_FIRST, SAL_COL_SAL_START + 2, 50.0)

    # Contract line 2 — partial plan (sum 75%)
    r2 = SAL_ROW_FIRST + 1
    sal.cell(r2, SAL_COL_ITEM, "SAL-VOCE-02")
    sal.cell(r2, SAL_COL_DESC, "Extra oneri / DOP — seconda voce")
    sal.cell(r2, SAL_COL_QTY, 1)
    sal.cell(r2, SAL_COL_UNIT_PRICE, 25000.0)
    sal.cell(r2, SAL_COL_SAL_START, 25.0)
    sal.cell(r2, SAL_COL_SAL_START + 1, 25.0)
    sal.cell(r2, SAL_COL_SAL_START + 2, 25.0)

    # OFFERTA — two data rows (used only if ANACO empty; still good for fallback drill)
    off = wb.create_sheet("OFFERTA")
    off.cell(23, OFFERTA_COL_CODE, "CODICE")
    off.cell(23, OFFERTA_COL_DESC, "DESCRIZIONE")
    off.cell(24, OFFERTA_COL_CODE, "FT-OFF-99")
    off.cell(24, OFFERTA_COL_WIDTH, 900)
    off.cell(24, OFFERTA_COL_HEIGHT, 2100)
    off.cell(24, OFFERTA_COL_DESC, "Riga OFFERTA fallback A")
    off.cell(24, OFFERTA_COL_UOM, "MQ")
    off.cell(24, OFFERTA_COL_QTY, 1)
    off.cell(24, OFFERTA_COL_PRICE, 1250.0)
    off.cell(25, OFFERTA_COL_CODE, "VC-OFF-88")
    off.cell(25, OFFERTA_COL_DESC, "Riga OFFERTA fallback B (nr)")
    off.cell(25, OFFERTA_COL_UOM, "Nr")
    off.cell(25, OFFERTA_COL_QTY, 4)
    off.cell(25, OFFERTA_COL_PRICE, 320.0)

    wb.save(out_path)
    print(f"Written: {out_path}")
    print(
        "Rich content: ANACO row5 K/L/M=0.95/0.96/0.97, BM=6% BP=4%; "
        "3 item rows (FT, LA, ACC) with CAD costs + vetro/accessori; "
        "2 SAL contract rows (100% + 75%); OFFERTA 2 fallback lines."
    )
    print(
        "Expected import: 3 ANACO lines, 2 SAL lines. "
        "Odoo should show non-zero Costo totale and realistic margin vs prima."
    )


if __name__ == "__main__":
    main()
