# -*- coding: utf-8 -*-
"""Unified TMS Excel import on project (RDA/ACO/ACP/LDS/drawings/VdC catalog)."""
import base64
import io

from odoo import _, api, fields, models
from odoo.exceptions import UserError

from .sbu_tms_excel_parser import (
    detect_tms_file_kind,
    parse_drawing_sheet_rows,
    parse_elenco_elementi_rows,
    parse_lds_sheet_rows,
    parse_tms_worksheet_rows,
    parse_vdc_sheet_rows,
    pick_tms_sheet,
)
from ..models.sbu_tms_helpers import resolve_tms_uom

try:
    import openpyxl
except ImportError:
    openpyxl = None


class SbuProjectTmsImportWizard(models.TransientModel):
    _name = 'sbu.project.tms.import.wizard'
    _description = 'Import TMS Excel files on project'

    project_id = fields.Many2one(
        'project.project',
        string='Project / Job',
        required=True,
        ondelete='cascade',
    )
    data_file = fields.Binary(string='Excel file', required=True)
    filename = fields.Char(string='Filename')
    import_kind = fields.Selection(
        [
            ('auto', 'Auto-detect from file'),
            ('rda', 'RDA — materials'),
            ('aco', 'ACO — workshop accessories'),
            ('acp', 'ACP — installation accessories'),
            ('lds', 'LDS register'),
            ('drawings', 'Drawing approval register'),
            ('vdc', 'VdC budget catalog only'),
            ('elenco', 'Elenco elementi (route owners)'),
        ],
        string='Import type',
        default='auto',
        required=True,
    )
    update_mode = fields.Selection(
        [
            ('merge', 'Update matching / add new'),
            ('replace', 'Replace all (PR lines or register rows)'),
        ],
        default='merge',
        required=True,
    )
    sync_vdc_catalog = fields.Boolean(
        string='Sync VdC catalog from «Vdc» sheet',
        default=True,
        help='When importing RDA workbooks, upsert budget codes from the Vdc worksheet.',
    )
    sync_elenco_routes = fields.Boolean(
        string='Sync route owners from «Elenco elementi»',
        default=True,
        help='Update workflow route default drawn-by from Elenco elementi sheet if present.',
    )
    create_purchase_request = fields.Boolean(
        string='Create purchase request if missing',
        default=True,
    )
    result_summary = fields.Text(string='Result', readonly=True)

    def _load_workbook(self):
        self.ensure_one()
        if not openpyxl:
            raise UserError(_('Install the openpyxl Python library on the Odoo server.'))
        from odoo.addons.sbu_estimate.wizards.sbu_openpyxl_utils import load_openpyxl_workbook

        raw = base64.b64decode(self.data_file)
        return load_openpyxl_workbook(raw, data_only=True, read_only=True)

    def _resolve_kind(self, wb):
        if self.import_kind != 'auto':
            return self.import_kind
        return detect_tms_file_kind(self.filename or '', wb.sheetnames)

    def _sheet_rows(self, wb, sheet_name):
        if sheet_name not in wb.sheetnames:
            return []
        return list(wb[sheet_name].iter_rows(values_only=True))

    def _find_or_create_pr(self, request_type):
        self.ensure_one()
        PurchaseRequest = self.env['sbu.purchase.request']
        pr = PurchaseRequest.search([
            ('project_id', '=', self.project_id.id),
            ('request_type', '=', request_type),
            ('state', 'not in', ('cancelled', 'done')),
        ], limit=1)
        if pr or not self.create_purchase_request:
            return pr
        route = self.env['sbu.workflow.route'].search([
            ('request_type', '=', request_type),
            ('wizard_enabled', '=', True),
        ], limit=1)
        defaults = self.env['sbu.workflow.route'].defaults_for_code(route.code) if route else {}
        return PurchaseRequest.create({
            'project_id': self.project_id.id,
            'request_type': request_type,
            'workflow_route': route.code if route else False,
            'company_id': self.project_id.company_id.id,
            'drawn_by': defaults.get('drawn_by') or False,
        })

    def _import_pr_document(self, wb, request_type):
        ws = pick_tms_sheet(wb, request_type, None)
        rows = list(ws.iter_rows(values_only=True))
        header, line_rows = parse_tms_worksheet_rows(rows)
        if not line_rows:
            raise UserError(_('No lines found in the TMS worksheet.'))
        pr = self._find_or_create_pr(request_type)
        if not pr:
            raise UserError(
                _('No open %(typ)s document on this job. Enable «Create purchase request» or create one first.')
                % {'typ': request_type.upper()},
            )
        wizard = self.env['sbu.purchase.request.excel.import.wizard'].create({
            'request_id': pr.id,
            'data_file': self.data_file,
            'filename': self.filename,
            'template_format': 'tms',
            'update_mode': self.update_mode,
            'import_header': True,
        })
        wizard.action_import()
        return pr, len(line_rows)

    def _sync_workbook_extras(self, wb):
        parts = []
        if self.sync_vdc_catalog and 'Vdc' in wb.sheetnames:
            vdc_rows = parse_vdc_sheet_rows(self._sheet_rows(wb, 'Vdc'))
            stats = self.env['sbu.vdc.catalog'].sync_from_sheet_rows(vdc_rows)
            parts.append(
                _('VdC catalog: %(created)s created, %(updated)s updated.',
                  created=stats['created'], updated=stats['updated']),
            )
        elenco_sheet = next(
            (n for n in wb.sheetnames if 'elenco elementi' in n.lower()),
            None,
        )
        if self.sync_elenco_routes and elenco_sheet:
            elenco_rows = parse_elenco_elementi_rows(self._sheet_rows(wb, elenco_sheet))
            n = self.env['sbu.workflow.route'].sync_elenco_from_rows(elenco_rows)
            parts.append(_('Route owners updated: %s.') % n)
        return parts

    def action_import(self):
        self.ensure_one()
        wb = self._load_workbook()
        try:
            kind = self._resolve_kind(wb)
            parts = []
            if kind in ('rda', 'aco', 'acp'):
                pr, n_lines = self._import_pr_document(wb, kind)
                parts.append(
                    _('%(typ)s → %(ref)s (%(n)s lines).',
                      typ=kind.upper(), ref=pr.display_name, n=n_lines),
                )
                parts.extend(self._sync_workbook_extras(wb))
            elif kind == 'lds':
                sheet = 'reg_LDS' if 'reg_LDS' in wb.sheetnames else wb.sheetnames[0]
                rows = parse_lds_sheet_rows(self._sheet_rows(wb, sheet))
                if not rows:
                    raise UserError(_('No LDS rows found.'))
                if self.update_mode == 'replace':
                    self.env['sbu.lds.entry'].search([
                        ('project_id', '=', self.project_id.id),
                    ]).unlink()
                stats = self.env['sbu.lds.entry'].import_tms_rows(
                    self.project_id, rows, self.update_mode,
                )
                parts.append(
                    _('LDS register: %(created)s created, %(updated)s updated.',
                      created=stats['created'], updated=stats['updated']),
                )
                parts.extend(self._sync_workbook_extras(wb))
            elif kind == 'drawings':
                sheet = next(
                    (n for n in wb.sheetnames if 'fac' in n.lower() or 'tav' in n.lower()),
                    wb.sheetnames[0],
                )
                rows = parse_drawing_sheet_rows(self._sheet_rows(wb, sheet))
                if not rows:
                    raise UserError(_('No drawing rows found.'))
                if self.update_mode == 'replace':
                    self.env['sbu.drawing.register'].search([
                        ('project_id', '=', self.project_id.id),
                    ]).unlink()
                stats = self.env['sbu.drawing.register'].import_tms_rows(
                    self.project_id, rows, self.update_mode,
                )
                parts.append(
                    _('Drawings: %(created)s created, %(updated)s updated.',
                      created=stats['created'], updated=stats['updated']),
                )
            elif kind == 'vdc':
                if 'Vdc' not in wb.sheetnames:
                    raise UserError(_('Worksheet «Vdc» not found.'))
                stats = self.env['sbu.vdc.catalog'].sync_from_sheet_rows(
                    parse_vdc_sheet_rows(self._sheet_rows(wb, 'Vdc')),
                )
                parts.append(
                    _('VdC catalog: %(created)s created, %(updated)s updated.',
                      created=stats['created'], updated=stats['updated']),
                )
            elif kind == 'elenco':
                sheet = next(
                    (n for n in wb.sheetnames if 'elenco' in n.lower()),
                    wb.sheetnames[0],
                )
                n = self.env['sbu.workflow.route'].sync_elenco_from_rows(
                    parse_elenco_elementi_rows(self._sheet_rows(wb, sheet)),
                )
                parts.append(_('Route owners updated: %s.') % n)
            else:
                raise UserError(_('Unsupported import type: %s') % kind)
            summary = '\n'.join(parts)
            self.result_summary = summary
            self.project_id.message_post(body=_('TMS import (%(kind)s): %(msg)s') % {
                'kind': kind,
                'msg': summary.replace('\n', ' · '),
            })
        finally:
            wb.close()
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'sbu.project.tms.import.wizard',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }
