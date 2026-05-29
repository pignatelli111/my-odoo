# -*- coding: utf-8 -*-
"""Parse Suburban TMS Excel templates (M.4.3.x RDA/ACO/ACP, M.4.4.x LDS).

Layout reference: ``file base che usano i tecnici_riservato/`` in the repo.
"""
from __future__ import annotations

import re
from datetime import date, datetime

# Preferred worksheet per Odoo request_type
TMS_SHEETS_BY_REQUEST_TYPE = {
    'rda': ('Esempio RDA', 'RDA'),
    'aco': ('Esempio ACO', 'ACO'),
    'acp': ('Esempio ACP', 'ACP'),
    'lds': ('reg_LDS', 'LDS'),
}

_HEADER_ALIASES = {
    'pos': ('pos', 'pos.', 'posizione', 'position'),
    'article_code': (
        'cod. articolo', 'cod articolo', 'codice articolo', 'article', 'cod.articolo',
    ),
    'name': ('descrizione', 'description', 'desc', 'nome'),
    'width_mm': ('l (mm)', 'l', 'larghezza', 'width', 'width mm', 'l mm'),
    'height_mm': ('h (mm)', 'h', 'altezza', 'height', 'height mm', 'h mm'),
    'depth_mm': ('p (mm)', 'p', 'profondità', 'profondita', 'depth', 'depth mm', 'p mm'),
    'sqm_per_piece': ('mq/cad', 'mq cad', 'mq per pezzo', 'sqm/cad'),
    'sqm_total': ('area (mq)', 'area mq', 'mq tot', 'mq tot.', 'mq totali', 'mq totale'),
    'product_qty': (
        'qty', 'quantità', 'quantita', 'q.tà', 'qta', 'quantity',
        'quantita tot', 'quantità tot', 'quantita totale',
    ),
    'product_uom': ('unità', 'unita', 'u.m.', 'um', 'uom'),
    'utilization': ('utilizzo', 'utilization', 'uso'),
    'vdc_code': ('vdc', 'voce di budget', 'codice vdc', 'codice vdc/vob'),
    'weight_kg': ('peso (tot) kg', 'peso tot kg', 'peso kg', 'peso (tot)', 'peso'),
    'date_required': ('data consegna', 'delivery date'),
    'reference_rda': ('riferimento rda', 'rif rda', 'ref rda'),
    'order_number': ('n° ordine', 'n ordine', 'n.ro ordine', 'numero ordine', 'n.ro                   ordine'),
    'note': ('note conformita', 'note conformità', 'note'),
    'destination': (
        'destination / destinazione', 'destinazione', 'destination',
    ),
    'logistics_supplier': (
        'fornitore/ logistica', 'fornitore/logistica', 'ddt consegne',
    ),
    'logistics_site': (
        'logistica/ cantiere', 'logistica/cantiere', 'logistica/             cantiere',
    ),
    'procurement_wh': ('magazzino', 'warehouse', 'approvvigionamento'),
}

_DIM_L_RE = re.compile(r'L\s*=?\s*([\d.,]+)\s*(mm|ml|m)?', re.IGNORECASE)
_DIM_H_RE = re.compile(r'H\s*=?\s*([\d.,]+)\s*(mm|ml|m)?', re.IGNORECASE)


def _norm_header(cell) -> str:
    if cell is None:
        return ''
    text = re.sub(r'\s+', ' ', str(cell).strip().lower())
    return text.replace("'", '').replace('’', '')


def _cell_str(row, idx) -> str:
    if idx is None or idx >= len(row):
        return ''
    val = row[idx]
    if val is None:
        return ''
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    return str(val).strip()


def _cell_float(row, idx) -> float:
    if idx is None or idx >= len(row):
        return 0.0
    val = row[idx]
    if val is None or val == '':
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    text = str(val).strip().replace(',', '.')
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def _parse_dim_text(text: str) -> tuple[float, float, float]:
    """Return (width_mm, height_mm, depth_mm) parsed from free-text dimension cells."""
    if not text:
        return 0.0, 0.0, 0.0
    w = h = p = 0.0
    m = _DIM_L_RE.search(text)
    if m:
        w = _parse_measure(m.group(1), m.group(2))
    m = _DIM_H_RE.search(text)
    if m:
        h = _parse_measure(m.group(1), m.group(2))
    return w, h, p


def _parse_measure(num_str: str, unit: str | None) -> float:
    try:
        val = float(num_str.replace(',', '.'))
    except ValueError:
        return 0.0
    unit = (unit or 'mm').lower()
    if unit == 'm':
        return val * 1000.0
    return val


def _map_headers(row_cells) -> dict:
    mapping = {}
    for idx, cell in enumerate(row_cells):
        h = _norm_header(cell)
        if not h:
            continue
        for field_name, aliases in _HEADER_ALIASES.items():
            if h in aliases or any(h.startswith(a) for a in aliases if len(a) > 4):
                mapping.setdefault(field_name, idx)
                break
            if field_name == 'product_qty' and 'quantit' in h and 'tot' in h:
                mapping.setdefault(field_name, idx)
                break
            if field_name == 'vdc_code' and 'vdc' in h:
                mapping.setdefault(field_name, idx)
                break
            if field_name == 'sqm_total' and 'area' in h and 'mq' in h:
                mapping.setdefault(field_name, idx)
                break
            if field_name == 'destination' and 'destinazione' in h:
                mapping.setdefault(field_name, idx)
                break
            if field_name == 'logistics_supplier' and 'fornitore' in h and 'logistica' in h:
                mapping.setdefault(field_name, idx)
                break
            if field_name == 'logistics_site' and 'logistica' in h and 'cantiere' in h:
                mapping.setdefault(field_name, idx)
                break
            if field_name == 'order_number' and 'ordine' in h:
                mapping.setdefault(field_name, idx)
                break
    return mapping


def _merge_subheader_row(mapping: dict, sub_row) -> None:
    for idx, cell in enumerate(sub_row):
        h = _norm_header(cell)
        if h in ('l', 'larghezza', 'width'):
            mapping.setdefault('width_mm', idx)
        elif h in ('h', 'altezza', 'height'):
            mapping.setdefault('height_mm', idx)
        elif h in ('p', 'profondita', 'profondità', 'depth'):
            mapping.setdefault('depth_mm', idx)


def is_tms_workbook(rows: list[tuple]) -> bool:
    """Detect Suburban TMS template from first ~20 rows."""
    for row in rows[:20]:
        joined = ' '.join(_norm_header(c) for c in row if c)
        if any(k in joined for k in (
            'richiesta di acquisto materiali',
            'accessori officina',
            'accessori posa',
            'riepilogo lds',
        )):
            return True
        if 'pos.' in joined and 'descrizione' in joined:
            if 'cod. articolo' in joined or 'cod articolo' in joined:
                return True
    return False


def pick_tms_sheet(wb, request_type: str, sheet_name: str | None):
    """Return worksheet for TMS import (prefer example sheet)."""
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f'Sheet «{sheet_name}» not found.')
        return wb[sheet_name]
    for candidate in TMS_SHEETS_BY_REQUEST_TYPE.get(request_type, ()):
        if candidate in wb.sheetnames:
            return wb[candidate]
    for name in wb.sheetnames:
        upper = name.upper()
        if request_type == 'rda' and 'RDA' in upper:
            return wb[name]
        if request_type == 'aco' and 'ACO' in upper:
            return wb[name]
        if request_type == 'acp' and 'ACP' in upper:
            return wb[name]
    return wb[wb.sheetnames[0]]


def _find_label_value(rows, labels: tuple[str, ...], max_row: int = 12) -> str:
    for row in rows[:max_row]:
        for c, cell in enumerate(row):
            text = _norm_header(cell)
            if not text:
                continue
            if any(lab in text for lab in labels):
                for off in range(1, 6):
                    if c + off < len(row) and row[c + off] not in (None, ''):
                        return _cell_str(row, c + off)
    return ''


def _parse_excel_date(val) -> date | None:
    if val is None or val == '':
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, (int, float)):
        try:
            from openpyxl.utils.datetime import from_excel
            dt = from_excel(val)
            return dt.date() if isinstance(dt, datetime) else dt
        except Exception:
            return None
    text = str(val).strip()
    for fmt in ('%d/%m/%Y', '%Y-%m-%d'):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def parse_tms_header(rows: list[tuple]) -> dict:
    """Extract RDA/ACO/ACP header block (project, item, topic, signatures, dates)."""
    header = {
        'excel_item': _find_label_value(rows, ('item',)),
        'topic': _find_label_value(rows, ('topic / argomento', 'topic', 'argomento')),
        'drawn_by': _find_label_value(rows, ('drawn by', 'redatto')),
        'check_by': _find_label_value(rows, ('check by', 'verificato')),
        'project_code': _find_label_value(rows, (
            'project / progetto', 'codice commessa', 'project',
        )),
        'document_date': _parse_excel_date(_find_label_value(rows, ('data / date', 'data'))),
    }
    for row in rows[6:10]:
        for c, cell in enumerate(row):
            h = _norm_header(cell)
            if h == 'item':
                for off in range(1, 8):
                    if c + off >= len(row):
                        break
                    val = _cell_str(row, c + off)
                    if not val or 'topic' in _norm_header(val):
                        continue
                    if not header.get('excel_item'):
                        header['excel_item'] = val
                    elif (
                        not header.get('area_code')
                        and len(val) <= 12
                        and val != header.get('excel_item')
                    ):
                        header['area_code'] = val
            if 'topic' in h and 'argomento' in h:
                for off in range(1, 8):
                    if c + off >= len(row):
                        break
                    val = _cell_str(row, c + off)
                    if val and 'topic' not in _norm_header(val):
                        header['topic'] = val
                        break
    for row in rows[:8]:
        joined = ' '.join(_norm_header(c) for c in row if c)
        if 'leed' in joined and not header.get('leed_scheme'):
            header['leed_scheme'] = 'LEED'
        if 'bream' in joined:
            header['leed_scheme'] = 'BREAM'
    for row in rows[6:9]:
        doc_num = _cell_str(row, 0)
        doc_label = _norm_header(doc_num)
        if doc_num and not doc_label.startswith('n°') and not doc_label.startswith('n '):
            header['tms_document_number'] = doc_num
            break
        template_name = _cell_str(row, 1)
        if template_name and 'tms_' in template_name.lower():
            header['tms_template_name'] = template_name
    need_by = None
    for row in rows[:8]:
        for c, cell in enumerate(row):
            if 'prima consegna' in _norm_header(cell) or 'prima cons' in _norm_header(cell):
                for off in range(1, 4):
                    d = _parse_excel_date(row[c + off] if c + off < len(row) else None)
                    if d:
                        need_by = d
                        break
        if need_by:
            break
    if not need_by:
        need_by = _parse_excel_date(_find_label_value(rows, ('data / date', 'data')))
    header['need_by_date'] = need_by
    header.update(parse_tms_delivery_dates(rows))
    return header


def parse_tms_delivery_dates(rows: list[tuple]) -> dict:
    """PRIMA / SECONDA / TERZA / QUARTA CONSEGNA from TMS header rows."""
    labels = (
        ('delivery_date_1', ('prima consegna', 'prima cons')),
        ('delivery_date_2', ('seconda consegna', 'seconda cons')),
        ('delivery_date_3', ('terza consegna', 'terza cons')),
        ('delivery_date_4', ('quarta consegna', 'quarta cons')),
    )
    out = {}
    for row_idx, row in enumerate(rows[:8]):
        for c, cell in enumerate(row):
            h = _norm_header(cell)
            for field, keys in labels:
                if any(k in h for k in keys):
                    val = None
                    for off in (1, 0):
                        target_row = rows[row_idx + off] if row_idx + off < len(rows) else ()
                        if c < len(target_row):
                            val = _parse_excel_date(target_row[c])
                        if val:
                            break
                    if val:
                        out[field] = val
    if out.get('delivery_date_1') and not out.get('need_by_date'):
        out['need_by_date'] = out['delivery_date_1']
    return out


def find_line_header(rows: list[tuple]) -> tuple[int, dict]:
    for i, row in enumerate(rows[:40]):
        mapping = _map_headers(row)
        if 'name' in mapping or ('pos' in mapping and 'article_code' in mapping):
            if i + 1 < len(rows):
                _merge_subheader_row(mapping, rows[i + 1])
            return i, mapping
    return -1, {}


def parse_tms_line_row(row, header_row: dict) -> dict | None:
    name = _cell_str(row, header_row.get('name'))
    pos = _cell_str(row, header_row.get('pos'))
    article = _cell_str(row, header_row.get('article_code'))
    if not name and not article and not pos:
        return None
    if _norm_header(name) in ('descrizione',) or _norm_header(pos) in ('pos.', 'pos'):
        return None

    width = _cell_float(row, header_row.get('width_mm'))
    height = _cell_float(row, header_row.get('height_mm'))
    depth = _cell_float(row, header_row.get('depth_mm'))

    dim_col = header_row.get('width_mm')
    if not width and not height and dim_col is not None:
        dim_text = _cell_str(row, dim_col)
        if dim_text and not dim_text.replace('.', '').isdigit():
            dw, dh, _dp = _parse_dim_text(dim_text)
            width = width or dw
            height = height or dh
    for idx in range(len(row)):
        text = _cell_str(row, idx)
        if 'L=' in text.upper() or 'H=' in text.upper():
            dw, dh, dd = _parse_dim_text(text)
            width = width or dw
            height = height or dh
            depth = depth or dd

    qty = _cell_float(row, header_row.get('product_qty')) or 1.0
    sqm_pc = _cell_float(row, header_row.get('sqm_per_piece'))
    sqm_tot = _cell_float(row, header_row.get('sqm_total'))
    if not sqm_tot and sqm_pc and qty:
        sqm_tot = sqm_pc * qty

    procurement = 'purchase'
    wh_idx = header_row.get('procurement_wh')
    if wh_idx is not None:
        cell_val = _norm_header(row[wh_idx] if wh_idx < len(row) else '')
        if cell_val in ('magazzino', 'warehouse', 'x', 'si'):
            procurement = 'warehouse'
        elif cell_val in ('acquisto', 'purchase'):
            procurement = 'purchase'
    for idx, cell in enumerate(row):
        h = _norm_header(cell)
        if h == 'magazzino' or h == 'x' and idx == wh_idx:
            procurement = 'warehouse'
            break

    dest = _cell_str(row, header_row.get('destination'))
    if not dest:
        dest = _cell_str(row, header_row.get('logistics_site'))

    date_req = _parse_excel_date(
        row[header_row['date_required']] if header_row.get('date_required') is not None
        and header_row['date_required'] < len(row) else None
    )

    return {
        'pos': pos,
        'article_code': article,
        'name': name or article or pos,
        'width_mm': width,
        'height_mm': height,
        'depth_mm': depth,
        'sqm_per_piece': sqm_pc,
        'sqm_total': sqm_tot,
        'product_qty': qty,
        'product_uom_code': _cell_str(row, header_row.get('product_uom')),
        'utilization': _cell_str(row, header_row.get('utilization')),
        'vdc_code': _cell_str(row, header_row.get('vdc_code')),
        'weight_kg': _cell_float(row, header_row.get('weight_kg')),
        'reference_rda': _cell_str(row, header_row.get('reference_rda')),
        'order_number': _cell_str(row, header_row.get('order_number')),
        'destination': dest,
        'logistics_supplier': _cell_str(row, header_row.get('logistics_supplier')),
        'logistics_site': _cell_str(row, header_row.get('logistics_site')),
        'note': _cell_str(row, header_row.get('note')),
        'date_required': date_req,
        'procurement_mode': procurement,
    }


def parse_tms_worksheet_rows(rows: list[tuple]) -> tuple[dict, list[dict]]:
    """Return (header_dict, line_dicts) from worksheet rows."""
    header = parse_tms_header(rows)
    header_idx, header_map = find_line_header(rows)
    if header_idx < 0:
        return header, []
    lines = []
    for row in rows[header_idx + 2:]:
        if not any(c is not None and str(c).strip() for c in row):
            continue
        parsed = parse_tms_line_row(row, header_map)
        if parsed:
            lines.append(parsed)
    return header, lines


def parse_vdc_sheet_rows(rows: list[tuple]) -> list[dict]:
    """Parse «Vdc» worksheet (code col B, description col C, pdc col D/E, note col F)."""
    entries = []
    for row in rows[1:120]:
        seq = _cell_str(row, 0)
        code = _cell_str(row, 1)
        name = _cell_str(row, 2)
        pdc_code = _cell_str(row, 3)
        pdc_label = _cell_str(row, 4)
        note = _cell_str(row, 5)
        if not code or _norm_header(code).startswith('vdc'):
            continue
        if not seq and not name:
            continue
        entries.append({
            'code': code,
            'name': name or code,
            'pdc_code': pdc_code,
            'pdc_label': pdc_label,
            'note': note,
        })
    return entries


def parse_lds_sheet_rows(rows: list[tuple]) -> list[dict]:
    header_idx = -1
    for i, row in enumerate(rows[:20]):
        joined = ' '.join(_norm_header(c) for c in row if c)
        if 'numero' in joined and 'lds' in joined:
            header_idx = i
            break
    if header_idx < 0:
        header_idx, _header_map = find_line_header(rows)
    if header_idx < 0:
        return []
    ref_cols = {}
    header_row = rows[header_idx]
    sub_row = rows[header_idx + 1] if header_idx + 1 < len(rows) else ()
    for idx, cell in enumerate(header_row):
        h = _norm_header(cell)
        if h == 'rda' or 'documento' in h and 'riferimento' in h:
            ref_cols['reference_rda'] = idx
        elif h == 'acp':
            ref_cols['reference_acp'] = idx
        elif h == 'aco':
            ref_cols['reference_aco'] = idx
    for idx, cell in enumerate(sub_row):
        h = _norm_header(cell)
        if h == 'rda':
            ref_cols['reference_rda'] = idx
        elif h == 'acp':
            ref_cols['reference_acp'] = idx
        elif h == 'aco':
            ref_cols['reference_aco'] = idx
    entries = []
    data_start = header_idx + 2
    for row in rows[data_start:]:
        if not any(c is not None and str(c).strip() for c in row):
            continue
        lds_number = _cell_str(row, 0)
        pos = _cell_str(row, 2)
        desc = _cell_str(row, 9)
        if not lds_number and not pos and not desc:
            continue
        material_raw = _norm_header(row[13] if len(row) > 13 else '')
        entries.append({
            'lds_number': lds_number,
            'delivery_date': _parse_excel_date(row[1] if len(row) > 1 else None),
            'pos': pos,
            'route_type': _cell_str(row, 3),
            'product_type': _cell_str(row, 4),
            'reference_rda': _cell_str(row, ref_cols.get('reference_rda', 5)),
            'reference_acp': _cell_str(row, ref_cols.get('reference_acp', 6)),
            'reference_aco': _cell_str(row, ref_cols.get('reference_aco', 7)),
            'order_number': _cell_str(row, 8),
            'description': desc,
            'article_code': _cell_str(row, 10),
            'product_qty': _cell_float(row, 11) or 1.0,
            'product_uom_code': _cell_str(row, 12),
            'material_present': material_raw in ('si', 'yes', 'presente', 'x', 'presente'),
            'expected_date': _parse_excel_date(row[14] if len(row) > 14 else None),
            'note': _cell_str(row, 15),
        })
    return entries


def parse_drawing_sheet_rows(rows: list[tuple]) -> list[dict]:
    """Parse «Report tav appr TMS» / FAC TMS layout."""
    header_idx = -1
    for i, row in enumerate(rows[:15]):
        joined = ' '.join(_norm_header(c) for c in row if c)
        if 'codice tavola' in joined or ('descrizione' in joined and 'emiss' in joined):
            header_idx = i
            break
    if header_idx < 0:
        for i, row in enumerate(rows[:15]):
            if _cell_str(row, 8).upper().startswith('CODICE'):
                header_idx = i
                break
    if header_idx < 0:
        return []
    entries = []
    for row in rows[header_idx + 1:]:
        code = _cell_str(row, 8)
        desc = _cell_str(row, 9)
        if not code and not desc:
            continue
        if _norm_header(code) in ('codice tavola',):
            continue
        approval = _cell_str(row, 14).lower()
        state = 'draft'
        if approval in ('appr', 'approved', 'ok', 'si'):
            state = 'approved'
        elif approval:
            state = 'issued'
        entries.append({
            'prog': _cell_str(row, 1),
            'item': _cell_str(row, 3),
            'num': _cell_str(row, 5),
            'revision': _cell_str(row, 6),
            'reference': _cell_str(row, 7),
            'drawing_code': code,
            'name': desc or code,
            'emission_1_date': _parse_excel_date(row[10] if len(row) > 10 else None),
            'emission_2_date': _parse_excel_date(row[11] if len(row) > 11 else None),
            'emission_3_date': _parse_excel_date(row[12] if len(row) > 12 else None),
            'last_emission_date': _parse_excel_date(row[13] if len(row) > 13 else None),
            'approval_state': state,
            'note': _cell_str(row, 15),
        })
    return entries


def detect_tms_file_kind(filename: str, sheetnames: list[str]) -> str:
    name = (filename or '').lower()
    sheets = ' '.join(sheetnames).lower()
    if 'lds' in name or 'reg_lds' in sheets:
        return 'lds'
    if 'tav appr' in name or 'fac tms' in sheets:
        return 'drawings'
    if 'aco' in name or 'accessori officina' in name:
        return 'aco'
    if 'acp' in name or 'accessori posa' in name:
        return 'acp'
    if 'rda' in name or 'richieste d' in name:
        return 'rda'
    if 'vdc' in sheets and 'rda' in sheets:
        return 'rda'
    return 'rda'


def load_workbook_rows(wb, sheet_name=None):
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
    return list(ws.iter_rows(values_only=True))


def parse_elenco_elementi_rows(rows: list[tuple]) -> list[dict]:
    """Parse «Elenco elementi» (route owners from M.4.4.A LDS workbook)."""
    header_idx = -1
    for i, row in enumerate(rows[:10]):
        joined = ' '.join(_norm_header(c) for c in row if c)
        if 'elemento' in joined or 'elenco' in joined:
            header_idx = i
            break
    start = header_idx + 1 if header_idx >= 0 else 2
    entries = []
    for row in rows[start:]:
        item = _cell_str(row, 0)
        if not item or _norm_header(item) in ('elemento',):
            continue
        author = _cell_str(row, 5) or _cell_str(row, 4)
        entries.append({
            'item_code': item,
            'enabled': _norm_header(row[2] if len(row) > 2 else '') in ('si', 'yes', 'x'),
            'drawn_by': _cell_str(row, 4),
            'author': author,
        })
    return entries
