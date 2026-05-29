# -*- coding: utf-8 -*-
"""Import / update RDA lines from Excel template (Cosimo punto 1 — dati tecnici)."""
import base64
import io
import re

from odoo import _, api, fields, models
from odoo.exceptions import UserError

from .sbu_tms_excel_parser import (
    is_tms_workbook,
    parse_tms_worksheet_rows,
    parse_vdc_sheet_rows,
    pick_tms_sheet,
)
from ..models.sbu_tms_helpers import resolve_tms_uom

try:
    import openpyxl
except ImportError:
    openpyxl = None

_HEADER_ALIASES = {
    'pos': ('pos', 'pos.', 'posizione', 'position'),
    'article_code': (
        'cod. articolo', 'cod articolo', 'codice articolo', 'article', 'cod.articolo',
        'cod articolo', 'sku',
    ),
    'name': ('descrizione', 'description', 'desc', 'nome'),
    'width_mm': ('l (mm)', 'l', 'larghezza', 'width', 'width mm', 'l mm'),
    'height_mm': ('h (mm)', 'h', 'altezza', 'height', 'height mm', 'h mm'),
    'depth_mm': ('p (mm)', 'p', 'profondità', 'profondita', 'depth', 'depth mm', 'p mm'),
    'sqm_per_piece': ('mq/cad', 'mq cad', 'mq per pezzo', 'sqm/cad'),
    'sqm_total': ('mq tot', 'mq tot.', 'mq totali', 'mq totale', 'sqm total'),
    'product_qty': ('qty', 'quantità', 'quantita', 'q.tà', 'qta', 'quantity'),
    'utilization': ('utilizzo', 'utilization', 'uso'),
}


def _norm_header(cell):
    if cell is None:
        return ''
    text = re.sub(r'\s+', ' ', str(cell).strip().lower())
    return text.replace("'", '').replace('’', '')


def _map_headers(row_cells):
    """Return {field_name: column_index} from first row."""
    mapping = {}
    for idx, cell in enumerate(row_cells):
        h = _norm_header(cell)
        if not h:
            continue
        for field_name, aliases in _HEADER_ALIASES.items():
            if h in aliases or h.replace('.', '') in aliases:
                mapping.setdefault(field_name, idx)
                break
            if field_name == 'product_qty' and 'quantit' in h and 'tot' in h:
                mapping.setdefault(field_name, idx)
                break
    return mapping


def _cell_float(row, idx):
    if idx is None or idx >= len(row):
        return 0.0
    val = row[idx]
    if val is None or val == '':
        return 0.0
    try:
        return float(val)
    except (TypeError, ValueError):
        return 0.0


def _cell_str(row, idx):
    if idx is None or idx >= len(row):
        return ''
    val = row[idx]
    if val is None:
        return ''
    return str(val).strip()


class SbuPurchaseRequestExcelImportWizard(models.TransientModel):
    _name = 'sbu.purchase.request.excel.import.wizard'
    _description = 'Import RDA lines from Excel'

    request_id = fields.Many2one(
        'sbu.purchase.request',
        string='Purchase request',
        required=True,
        ondelete='cascade',
    )
    data_file = fields.Binary(string='Excel file', required=True)
    filename = fields.Char(string='Filename')
    sheet_name = fields.Char(
        string='Sheet name',
        help='Leave empty: TMS templates use «Esempio RDA/ACO/ACP» automatically.',
    )
    template_format = fields.Selection(
        [
            ('auto', 'Auto-detect (TMS or generic)'),
            ('tms', 'Suburban TMS (M.4.3.x)'),
            ('generic', 'Generic flat header'),
        ],
        string='Excel layout',
        default='auto',
        required=True,
    )
    update_mode = fields.Selection(
        [
            ('merge', 'Update matching lines / add new'),
            ('replace', 'Replace all lines from file'),
        ],
        string='Mode',
        default='merge',
        required=True,
    )
    import_header = fields.Boolean(
        string='Import header (Item, Topic, Drawn by, …)',
        default=True,
        help='When using TMS layout, fill request header from the Excel block.',
    )
    sync_vdc_catalog = fields.Boolean(
        string='Sync VdC catalog from «Vdc» sheet',
        default=True,
        help='Upsert budget codes from the Vdc worksheet when present in the workbook.',
    )
    line_count = fields.Integer(string='Lines processed', readonly=True)
    template_detected = fields.Char(string='Layout detected', readonly=True)

    def _load_workbook_rows(self):
        self.ensure_one()
        if not openpyxl:
            raise UserError(_('Install the openpyxl Python library on the Odoo server.'))
        raw = base64.b64decode(self.data_file)
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
        try:
            if self.template_format == 'generic':
                ws = wb[self.sheet_name] if self.sheet_name else wb[wb.sheetnames[0]]
            else:
                try:
                    ws = pick_tms_sheet(
                        wb,
                        self.request_id.request_type or 'rda',
                        self.sheet_name or None,
                    )
                except ValueError as err:
                    raise UserError(str(err)) from err
            rows = list(ws.iter_rows(values_only=True))
            if self.sync_vdc_catalog and 'Vdc' in wb.sheetnames:
                vdc_rows = parse_vdc_sheet_rows(
                    list(wb['Vdc'].iter_rows(values_only=True)),
                )
                self.env['sbu.vdc.catalog'].sync_from_sheet_rows(vdc_rows)
        finally:
            wb.close()
        if not rows:
            raise UserError(_('The worksheet is empty.'))
        return rows

    def _parse_generic_rows(self, rows):
        header_row = None
        header_idx = 0
        for i, row in enumerate(rows[:30]):
            mapping = _map_headers(row)
            if 'name' in mapping or ('pos' in mapping and 'article_code' in mapping):
                header_row = mapping
                header_idx = i
                break
        if not header_row:
            raise UserError(
                _('Could not find a header row. Expected columns such as POS, Descrizione, L, H, Qty.')
            )
        data_rows = []
        for row in rows[header_idx + 1:]:
            if not any(c is not None and str(c).strip() for c in row):
                continue
            name = _cell_str(row, header_row.get('name'))
            pos = _cell_str(row, header_row.get('pos'))
            article = _cell_str(row, header_row.get('article_code'))
            if not name and not article and not pos:
                continue
            data_rows.append({
                'pos': pos,
                'article_code': article,
                'name': name or article or _('Imported line'),
                'width_mm': _cell_float(row, header_row.get('width_mm')),
                'height_mm': _cell_float(row, header_row.get('height_mm')),
                'depth_mm': _cell_float(row, header_row.get('depth_mm')),
                'sqm_per_piece': _cell_float(row, header_row.get('sqm_per_piece')),
                'sqm_total': _cell_float(row, header_row.get('sqm_total')),
                'product_qty': _cell_float(row, header_row.get('product_qty')) or 1.0,
                'utilization': _cell_str(row, header_row.get('utilization')),
            })
        return {}, data_rows, 'generic'

    def _parse_workbook(self):
        rows = self._load_workbook_rows()
        use_tms = self.template_format == 'tms' or (
            self.template_format == 'auto' and is_tms_workbook(rows)
        )
        if use_tms:
            header, data_rows = parse_tms_worksheet_rows(rows)
            return header, data_rows, 'tms'
        return self._parse_generic_rows(rows)

    def _apply_header(self, header):
        if not self.import_header or not header:
            return
        vals = {}
        for key, field in (
            ('excel_item', 'excel_item'),
            ('topic', 'topic'),
            ('drawn_by', 'drawn_by'),
            ('check_by', 'check_by'),
            ('project_code', 'project_code'),
            ('tms_document_number', 'tms_document_number'),
            ('tms_template_name', 'tms_template_name'),
            ('document_date', 'document_date'),
            ('area_code', 'area_code'),
            ('need_by_date', 'need_by_date'),
            ('delivery_date_2', 'delivery_date_2'),
            ('delivery_date_3', 'delivery_date_3'),
            ('delivery_date_4', 'delivery_date_4'),
        ):
            if header.get(key):
                vals[field] = header[key]
        leed = (header.get('leed_scheme') or '').lower()
        if leed in ('leed', 'bream'):
            vals['leed_scheme'] = leed
        if vals:
            self.request_id.write(vals)

    def _match_line(self, Line, row):
        domain = [('request_id', '=', self.request_id.id)]
        if row.get('pos'):
            domain.append(('pos', '=', row['pos']))
        if row.get('article_code'):
            domain.append(('article_code', '=', row['article_code']))
        return Line.search(domain, limit=1)

    def _line_vals(self, row, product):
        vals = {
            'pos': row.get('pos') or False,
            'article_code': row.get('article_code') or False,
            'name': row['name'],
            'width_mm': row.get('width_mm') or 0.0,
            'height_mm': row.get('height_mm') or 0.0,
            'depth_mm': row.get('depth_mm') or 0.0,
            'sqm_per_piece': row.get('sqm_per_piece') or 0.0,
            'sqm_total': row.get('sqm_total') or 0.0,
            'product_qty': row.get('product_qty') or 1.0,
            'utilization': row.get('utilization') or False,
            'vdc_code': row.get('vdc_code') or False,
            'reference_rda': row.get('reference_rda') or False,
            'order_number': row.get('order_number') or False,
            'logistics_supplier': row.get('logistics_supplier') or False,
            'logistics_site': row.get('logistics_site') or False,
            'note': row.get('note') or False,
            'weight_kg': row.get('weight_kg') or 0.0,
            'destination': row.get('destination') or False,
            'procurement_mode': row.get('procurement_mode') or 'purchase',
            'line_priority': self.request_id.priority,
        }
        if row.get('date_required'):
            vals['date_required'] = row['date_required']
        uom_code = row.get('product_uom_code')
        if uom_code:
            uom = resolve_tms_uom(self.env, uom_code)
            if uom:
                vals['product_uom'] = uom.id
        if product:
            vals['product_id'] = product.id
            vals['product_uom'] = product.uom_id.id
        return vals

    def action_import(self):
        self.ensure_one()
        if self.request_id.state == 'cancelled':
            raise UserError(_('Cannot import lines on a cancelled request.'))
        header, rows, layout = self._parse_workbook()
        if not rows:
            raise UserError(_('No data rows found below the header.'))
        self._apply_header(header)
        Line = self.env['sbu.purchase.request.line']
        Product = self.env['product.product']
        if self.update_mode == 'replace':
            self.request_id.line_ids.unlink()
        created = updated = 0
        for row in rows:
            product = False
            if row.get('article_code'):
                product = Product.search(
                    [('default_code', '=', row['article_code'])],
                    limit=1,
                )
            vals = self._line_vals(row, product)
            existing = self._match_line(Line, row) if self.update_mode == 'merge' else Line.browse()
            if existing:
                existing.write(vals)
                updated += 1
            else:
                vals['request_id'] = self.request_id.id
                Line.create(vals)
                created += 1
        self.request_id.write({'technical_data_state': 'excel_imported'})
        self.template_detected = layout
        self.line_count = created + updated
        self.request_id.message_post(
            body=_(
                'Excel import (%(layout)s): %(c)s created, %(u)s updated (%(n)s rows).',
                layout=layout,
                c=created,
                u=updated,
                n=len(rows),
            ),
        )
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'sbu.purchase.request',
            'res_id': self.request_id.id,
            'view_mode': 'form',
            'target': 'current',
        }
