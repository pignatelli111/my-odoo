# -*- coding: utf-8 -*-
"""TMS Excel layout parser tests (M.4.3.x templates)."""
import base64
import io

from odoo.tests import tagged
from odoo.tests.common import TransactionCase

from odoo.addons.sbu_purchase_flow.wizards.sbu_tms_excel_parser import (
    is_tms_workbook,
    parse_tms_delivery_dates,
    parse_tms_worksheet_rows,
    parse_vdc_sheet_rows,
)

try:
    import openpyxl
except ImportError:
    openpyxl = None


def _tms_rda_rows():
    """Minimal TMS RDA block matching tecnici_riservato layout."""
    rows = [''] * 12
    grid = [[''] * 20 for _ in range(15)]
    grid[1][8] = 'RICHIESTA DI ACQUISTO MATERIALI - RDA'
    grid[1][13] = 'P0146_2024'
    grid[3][10] = 'Drawn by / Redatto  da'
    grid[3][13] = 'D. Capelli'
    grid[4][10] = 'Check by / Verificato Da'
    grid[4][13] = 'V. Rabino'
    grid[6][10] = 'Item'
    grid[6][11] = 'SER'
    grid[6][13] = 'TOPIC / ARGOMENTO'
    grid[6][15] = 'Spessori falsi telai'
    grid[10][0] = 'POS.'
    grid[10][1] = 'DESCRIZIONE'
    grid[10][5] = 'COD. ARTICOLO'
    grid[10][6] = "QUANTITA' TOT"
    grid[10][7] = 'Unità'
    grid[10][9] = 'CODICE VdC'
    grid[10][10] = 'UTILIZZO'
    grid[10][11] = 'DIMENSIONE  (mm)'
    grid[11][11] = 'L'
    grid[11][12] = 'h'
    grid[12][0] = '01'
    grid[12][1] = 'Spessore falso telaio PVC'
    grid[12][5] = 'SER-TAS.02A'
    grid[12][6] = 650
    grid[12][7] = 'PZ'
    grid[12][9] = '44'
    grid[12][10] = 'montante'
    grid[12][11] = 60
    grid[12][12] = 100
    return [tuple(r) for r in grid]


@tagged('post_install', '-at_install')
class TestSbuTmsExcelParser(TransactionCase):

    def test_detect_tms_layout(self):
        rows = _tms_rda_rows()
        self.assertTrue(is_tms_workbook(rows))

    def test_parse_tms_rda_block(self):
        header, lines = parse_tms_worksheet_rows(_tms_rda_rows())
        self.assertEqual(header['drawn_by'], 'D. Capelli')
        self.assertEqual(header['check_by'], 'V. Rabino')
        self.assertEqual(header['excel_item'], 'SER')
        self.assertEqual(header['topic'], 'Spessori falsi telai')
        self.assertEqual(len(lines), 1)
        line = lines[0]
        self.assertEqual(line['pos'], '01')
        self.assertEqual(line['article_code'], 'SER-TAS.02A')
        self.assertEqual(line['product_qty'], 650.0)
        self.assertEqual(line['width_mm'], 60.0)
        self.assertEqual(line['height_mm'], 100.0)
        self.assertEqual(line['utilization'], 'montante')
        self.assertEqual(line['vdc_code'], '44')

    def test_parse_aco_example_fields(self):
        grid = [[''] * 20 for _ in range(15)]
        grid[9][0] = 'POS.'
        grid[9][1] = 'DESCRIZIONE'
        grid[9][2] = 'COD. ARTICOLO'
        grid[9][3] = "QUANTITA' TOT"
        grid[9][4] = 'Unità'
        grid[9][7] = 'UTILIZZO'
        grid[9][12] = 'RIFERIMENTO RDA'
        grid[9][13] = 'DATA CONSEGNA'
        grid[9][14] = 'N° ORDINE'
        grid[9][15] = 'DESTINATION / DESTINAZIONE'
        grid[10][3] = 'L'
        grid[10][4] = 'H'
        grid[11][0] = '01'
        grid[11][1] = 'TAPPI BMP'
        grid[11][2] = '25483A'
        grid[11][3] = 50
        grid[11][4] = 'pz'
        grid[11][7] = 'montante'
        grid[11][12] = 'P0587_2018_RDA_077'
        grid[11][14] = 'N°2018/2255'
        grid[11][15] = 'OFFICINA DI ASSEMBLAGGIO'
        rows = [tuple(r) for r in grid]
        _, lines = parse_tms_worksheet_rows(rows)
        line = lines[0]
        self.assertEqual(line['article_code'], '25483A')
        self.assertEqual(line['reference_rda'], 'P0587_2018_RDA_077')
        self.assertEqual(line['order_number'], 'N°2018/2255')
        self.assertEqual(line['destination'], 'OFFICINA DI ASSEMBLAGGIO')
        self.assertEqual(line['product_uom_code'], 'pz')

    def test_parse_delivery_dates(self):
        grid = [[''] * 20 for _ in range(8)]
        grid[2][10] = 'PRIMA CONSEGNA'
        grid[2][12] = 'SECONDA CONSEGNA'
        grid[3][10] = '2026-06-01'
        grid[3][12] = '2026-07-15'
        rows = [tuple(r) for r in grid]
        dates = parse_tms_delivery_dates(rows)
        self.assertEqual(str(dates.get('delivery_date_1')), '2026-06-01')
        self.assertEqual(str(dates.get('delivery_date_2')), '2026-07-15')

    def test_vdc_catalog_sync(self):
        rows = [
            {'code': 'B_GAMMISTA_ALLUMINIO', 'name': 'Gammista alluminio', 'pdc_code': 'M1', 'pdc_label': 'Mat'},
        ]
        stats = self.env['sbu.vdc.catalog'].sync_from_sheet_rows(rows)
        self.assertEqual(stats['created'], 1)
        cat = self.env['sbu.vdc.catalog'].search([('code', '=', 'B_GAMMISTA_ALLUMINIO')], limit=1)
        self.assertTrue(cat)
        self.assertEqual(cat.cost_family, 'aluminum_sheet')
        fam = self.env['sbu.vdc.catalog'].resolve_cost_family('B_GAMMISTA_ALLUMINIO')
        self.assertEqual(fam, 'aluminum_sheet')


@tagged('post_install', '-at_install')
class TestSbuProjectTmsImportWizard(TransactionCase):

    def test_project_tms_import_vdc_only(self):
        if not openpyxl:
            self.skipTest('openpyxl not installed')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Vdc'
        ws.append(['', 'Vdc', 'Description', 'Pdc', 'Label'])
        ws.append(['1', 'E_TEST_CODE', 'Test extra', 'E1', 'Extra'])
        buf = io.BytesIO()
        wb.save(buf)
        partner = self.env['res.partner'].create({'name': 'TMS partner 2'})
        estimate = self.env['sbu.estimate'].create({'partner_id': partner.id})
        project = self.env['project.project'].create({
            'name': 'TMS project 2',
            'sbu_estimate_id': estimate.id,
        })
        wizard = self.env['sbu.project.tms.import.wizard'].create({
            'project_id': project.id,
            'data_file': base64.b64encode(buf.getvalue()),
            'filename': 'vdc_only.xlsx',
            'import_kind': 'vdc',
        })
        wizard.action_import()
        self.assertTrue(
            self.env['sbu.vdc.catalog'].search_count([('code', '=', 'E_TEST_CODE')]),
        )


@tagged('post_install', '-at_install')
class TestSbuTmsExcelImportWizard(TransactionCase):

    def _workbook_b64(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Esempio RDA'
        for r_idx, row in enumerate(_tms_rda_rows(), start=1):
            for c_idx, val in enumerate(row, start=1):
                if val:
                    ws.cell(row=r_idx, column=c_idx, value=val)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def test_wizard_imports_tms_example_rda(self):
        if not openpyxl:
            self.skipTest('openpyxl not installed')
        partner = self.env['res.partner'].create({'name': 'TMS partner'})
        estimate = self.env['sbu.estimate'].create({'partner_id': partner.id})
        project = self.env['project.project'].create({
            'name': 'TMS project',
            'sbu_estimate_id': estimate.id,
        })
        pr = self.env['sbu.purchase.request'].create({
            'project_id': project.id,
            'request_type': 'rda',
            'workflow_route': 'LA',
        })
        wizard = self.env['sbu.purchase.request.excel.import.wizard'].create({
            'request_id': pr.id,
            'data_file': base64.b64encode(self._workbook_b64()),
            'filename': 'M.4.3.C_TMS_RDA_LEED.xlsx',
            'template_format': 'tms',
            'update_mode': 'replace',
        })
        wizard.action_import()
        self.assertEqual(pr.drawn_by, 'D. Capelli')
        self.assertEqual(pr.excel_item, 'SER')
        self.assertEqual(len(pr.line_ids), 1)
        line = pr.line_ids
        self.assertEqual(line.width_mm, 60.0)
        self.assertEqual(line.height_mm, 100.0)
        self.assertEqual(line.vdc_code, '44')
        self.assertEqual(pr.technical_data_state, 'excel_imported')
